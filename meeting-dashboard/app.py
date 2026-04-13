"""
会議資料ダッシュボード - Flask バックエンド
BigQuery から実績データを取得、Firestore / Cloud Storage でデータを管理
"""
import os
# ── ローカル開発用: .env ファイルから環境変数を読み込む ────
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    with open(_env_path, encoding="utf-8") as _f:
        for _line in _f:
            _line = _line.strip()
            if not _line or _line.startswith("#") or "=" not in _line:
                continue
            _k, _v = _line.split("=", 1)
            _k, _v = _k.strip(), _v.strip().strip('"').strip("'")
            if _k and not os.environ.get(_k):
                os.environ[_k] = _v
    print(f".env ファイル読み込み済み")

import json
import uuid
import sqlite3
import secrets
import threading
from datetime import date, datetime, timedelta, timezone
from functools import wraps
from pathlib import Path


from flask import Flask, Response, jsonify, request, send_from_directory, session, redirect
from google.cloud import bigquery
import pandas as pd
import firebase_admin
from firebase_admin import credentials as fb_credentials, auth as fb_auth, firestore

from week_mapping import (
    get_week_info, get_weeks_for_month, get_week_date_range,
    get_months_range, get_3month_weeks, get_2month_weeks, month_label, MONTH_WEEK_MAP,
    load_from_firestore as _load_week_from_fs,
    refresh_from_gas as _refresh_week_from_gas,
    save_to_firestore as _save_week_to_fs,
    _apply_rows as _apply_week_rows,
)

# ── Gemini AI ─────────────────────────────────────────
try:
    from google import genai
    _gemini_available = True
except ImportError:
    _gemini_available = False

# ── Claude AI ─────────────────────────────────────────
try:
    import anthropic
    _claude_available = True
except ImportError:
    _claude_available = False

# ── 環境判定 ──────────────────────────────────────────
IS_CLOUD_RUN = bool(os.environ.get("K_SERVICE"))  # Cloud Run では自動設定される

BASE_DIR = Path(__file__).parent
BOOTSTRAP_ADMIN_EMAIL = "matsunaga@ekmtc.com"
FIRESTORE_USERS_COLLECTION = "meeting_users"
DB_PATH = BASE_DIR / "data" / "prospects.db"
UPLOAD_DIR = BASE_DIR / "data" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

PROJECT_ID = "booking-data-388605"
FIRESTORE_DB_ID = "firestore-database-1012359"
FIRESTORE_CONFIG_COLLECTION = "meeting_config"
REFRESH_DOC_ID = "bq_refresh"

# タイムゾーン
JST = timezone(timedelta(hours=9))

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB
# Cloud Run では環境変数から固定キーを取得（複数インスタンス対応）
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))

# ── Firebase Admin SDK 初期化 ─────────────────────────
_firebase_cred_json = os.environ.get("FIREBASE_CREDENTIALS_JSON")  # Cloud Run: Secret Manager経由
if _firebase_cred_json:
    # Cloud Run: 環境変数にJSON文字列
    _fb_cred = fb_credentials.Certificate(json.loads(_firebase_cred_json))
else:
    # ローカル開発: ファイルパス
    FIREBASE_SERVICE_ACCOUNT_KEY = Path(
        r"C:\Users\matsunaga\Claude-Code-Test\Freight-News\serviceAccountKey.json"
    )
    _fb_cred = fb_credentials.Certificate(str(FIREBASE_SERVICE_ACCOUNT_KEY))

firebase_admin.initialize_app(_fb_cred)
try:
    db = firestore.client(database_id=FIRESTORE_DB_ID)
    _firestore_available = True
    print("Firestore: 接続OK")
except Exception as _e:
    db = None
    _firestore_available = False
    print(f"Firestore: 利用不可 ({_e})")

# ── BigQuery クライアント ──────────────────────────────
if IS_CLOUD_RUN:
    # Cloud Run: Application Default Credentials (サービスアカウント自動認証)
    from google.auth import default as _google_auth_default
    _bq_credentials, _ = _google_auth_default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    bq_client = bigquery.Client(project=PROJECT_ID, credentials=_bq_credentials)
else:
    # ローカル開発: サービスアカウントキーファイル
    SERVICE_ACCOUNT_KEY = Path(
        r"C:\Users\matsunaga\Documents\key\booking-data-388605@appspot.gserviceaccount.com\booking-data-388605-ec9e7af2c0e1.json"
    )
    from google.oauth2 import service_account as _sa
    _bq_credentials = _sa.Credentials.from_service_account_file(
        str(SERVICE_ACCOUNT_KEY),
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    bq_client = bigquery.Client(project=PROJECT_ID, credentials=_bq_credentials)

# ── 週マッピングを Firestore キャッシュからロード ─────────
if db:
    _load_week_from_fs(db)
    # GAS URL を Firestore 設定から復元
    try:
        import week_mapping as _wm_init
        _wm_cfg = db.collection("meeting_config").document("week_mapping_config").get()
        if _wm_cfg.exists:
            _wm_init.GAS_WEEK_URL = _wm_cfg.to_dict().get("gas_url", "")
            if _wm_init.GAS_WEEK_URL:
                print(f"week_mapping: GAS URL 設定済み")
    except Exception:
        pass

# ── Google Cloud Storage 初期化 ──────────────────────
GCS_BUCKET_NAME = os.environ.get("GCS_BUCKET", "meeting-dashboard-uploads-388605")
_gcs_available = False
gcs_bucket = None

try:
    from google.cloud import storage as gcs_storage
    if IS_CLOUD_RUN:
        _gcs_client = gcs_storage.Client(project=PROJECT_ID)
    else:
        _gcs_client = gcs_storage.Client(project=PROJECT_ID, credentials=_bq_credentials)
    gcs_bucket = _gcs_client.bucket(GCS_BUCKET_NAME)
    if not gcs_bucket.exists():
        gcs_bucket = _gcs_client.create_bucket(GCS_BUCKET_NAME, location="asia-northeast1")
        print(f"GCS: バケット作成OK ({GCS_BUCKET_NAME})")
    else:
        print(f"GCS: バケット接続OK ({GCS_BUCKET_NAME})")
    _gcs_available = True
except Exception as _gcs_e:
    print(f"GCS: 利用不可 ({_gcs_e}) - ローカルファイルを使用")
    _gcs_available = False

# ── Gemini AI クライアント初期化 ──────────────────────
_gemini_client = None
if _gemini_available:
    try:
        _gemini_kwargs = dict(vertexai=True, project=PROJECT_ID, location="us-central1")
        if not IS_CLOUD_RUN:
            _gemini_kwargs["credentials"] = _bq_credentials
        _gemini_client = genai.Client(**_gemini_kwargs)
        print("Gemini AI: 接続OK")
    except Exception as _gem_e:
        print(f"Gemini AI: 利用不可 ({_gem_e})")

# ── Claude AI クライアント初期化 ──────────────────────
_claude_client = None
if _claude_available:
    try:
        _anthropic_api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if _anthropic_api_key:
            _claude_client = anthropic.Anthropic(api_key=_anthropic_api_key)
            print("Claude AI: 接続OK")
        else:
            print("Claude AI: ANTHROPIC_API_KEY が未設定")
    except Exception as _cl_e:
        print(f"Claude AI: 利用不可 ({_cl_e})")

# Firestore コレクション名
FS_PROSPECTS = "meeting_prospects"
FS_MONTHLY_PROSPECTS = "meeting_monthly_prospects"
FS_NOTES = "meeting_notes"
FS_IMAGES = "meeting_images"
FS_BLOCKS = "meeting_blocks"
FS_SNAPSHOTS = "meeting_snapshots"
FS_TEMPLATE_CONFIG = "meeting_template_config"

# ── 親エリア自動合算: 子エリアの見込みを合算して親に反映 ──
AREA_AUTO_SUM = {
    "KR": ["JPC_KR", "JPN_KR"],
    "PH": ["MNL", "MIP"],
}
# 逆引き: 子エリア → 親エリア
_CHILD_TO_PARENT = {}
for _parent, _children in AREA_AUTO_SUM.items():
    for _child in _children:
        _CHILD_TO_PARENT[_child] = _parent

# テンプレート一覧 (デフォルト表示順)
TEMPLATE_DEFINITIONS = [
    {"id": "shipper_increase_curr", "label": "📈 増加荷主 TOP3（当月）", "default_on": True,  "col_span": 1},
    {"id": "shipper_increase_next", "label": "📈 増加荷主 TOP3（翌月）", "default_on": True,  "col_span": 1},
    {"id": "shipper_decrease_curr", "label": "📉 減少荷主 TOP3（当月）", "default_on": True,  "col_span": 1},
    {"id": "shipper_decrease_next", "label": "📉 減少荷主 TOP3（翌月）", "default_on": True,  "col_span": 1},
    {"id": "combo_increase_curr",  "label": "📈 荷主×航路 増加 TOP5（当月）", "default_on": True,  "col_span": 1},
    {"id": "combo_increase_next",  "label": "📈 荷主×航路 増加 TOP5（翌月）", "default_on": True,  "col_span": 1},
    {"id": "combo_decrease_curr",  "label": "📉 荷主×航路 減少 TOP5（当月）", "default_on": True,  "col_span": 1},
    {"id": "combo_decrease_next",  "label": "📉 荷主×航路 減少 TOP5（翌月）", "default_on": True,  "col_span": 1},
    {"id": "cm1_range",          "label": "💰 CM1レンジ分析",       "default_on": True,  "col_span": 1},
    {"id": "new_customer",       "label": "🆕 New Customer",        "default_on": True,  "col_span": 1},
    {"id": "regain_customer",    "label": "🔄 Regain Customer",     "default_on": True,  "col_span": 1},
    {"id": "trade_lane",         "label": "🗺️ Trade Lane",          "default_on": True,  "col_span": 2},
    {"id": "cm1_waterfall",      "label": "📊 CM1/TEU 要因分解",    "default_on": False, "col_span": 1},
    {"id": "booking_monthly",    "label": "📋 Booking件数（月間）",  "default_on": True, "col_span": 1},
    {"id": "booking_weekly",     "label": "📋 Booking件数（週間）",  "default_on": True, "col_span": 1},
    {"id": "pol_count",          "label": "🏭 POL数",               "default_on": True, "col_span": 1},
    {"id": "sales_contribution", "label": "👤 営業マン寄与度",      "default_on": False, "col_span": 1},
    {"id": "koshi_shipper",      "label": "📦 古紙荷主",             "default_on": True,  "col_span": 1},
]

# ── SQLite 初期化 (フォールバック用) ─────────────────────
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 見込みデータ (週 × エリア)
    c.execute("""
        CREATE TABLE IF NOT EXISTS prospects (
            week_key TEXT NOT NULL,   -- "2026-W09"
            area     TEXT NOT NULL,   -- "KR"
            teu      REAL,
            cm1      REAL,
            updated  TEXT,
            PRIMARY KEY (week_key, area)
        )
    """)
    # エリアメモ (週 × エリア)
    c.execute("""
        CREATE TABLE IF NOT EXISTS notes (
            week_key TEXT NOT NULL,
            area     TEXT NOT NULL,
            note     TEXT,
            updated  TEXT,
            PRIMARY KEY (week_key, area)
        )
    """)
    # 画像メタデータ
    c.execute("""
        CREATE TABLE IF NOT EXISTS images (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            week_key TEXT NOT NULL,
            area     TEXT NOT NULL,
            filename TEXT NOT NULL,
            caption  TEXT,
            uploaded TEXT
        )
    """)
    # 月間見込み (月 × エリア)
    c.execute("""
        CREATE TABLE IF NOT EXISTS monthly_prospects (
            ym          TEXT NOT NULL,
            area        TEXT NOT NULL,
            teu         REAL,
            cm1_per_teu REAL,
            updated     TEXT,
            PRIMARY KEY (ym, area)
        )
    """)
    # ブロックエディター (週 × エリア × 順序)
    c.execute("""
        CREATE TABLE IF NOT EXISTS blocks (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            week_key    TEXT NOT NULL,
            area        TEXT NOT NULL,
            block_order INTEGER NOT NULL DEFAULT 0,
            block_type  TEXT NOT NULL,
            content     TEXT,
            filename    TEXT,
            img_width   INTEGER DEFAULT 200,
            updated     TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ── Firestore ユーザー管理 ────────────────────────────
EDITOR_TEAMS = {"jpDigitalStrategy", "jpDigitalSales"}

def get_fs_user(email: str) -> dict | None:
    """Firestoreからユーザー情報を取得。team フィールドでロールを判定"""
    doc = db.collection(FIRESTORE_USERS_COLLECTION).document(email).get()
    if doc.exists:
        d = doc.to_dict()
        team = d.get("team", "")
        # team フィールドが editor_teams にあれば editor、なければ viewer
        if team in EDITOR_TEAMS:
            role = "editor"
        else:
            role = d.get("role", "viewer")  # 後方互換: role フィールドにフォールバック
        return {"email": email, "role": role,
                "display_name": d.get("display_name", email),
                "team": team}
    return None

def ensure_bootstrap_admin():
    """初期管理者を登録 / team フィールドが未設定なら追加"""
    col = db.collection(FIRESTORE_USERS_COLLECTION)
    doc_ref = col.document(BOOTSTRAP_ADMIN_EMAIL)
    doc = doc_ref.get()
    if not doc.exists:
        doc_ref.set({
            "team": "jpDigitalStrategy",
            "display_name": "管理者",
            "created_by": "system",
        })
        print(f"Bootstrap admin created: {BOOTSTRAP_ADMIN_EMAIL}")
    elif not doc.to_dict().get("team"):
        doc_ref.update({"team": "jpDigitalStrategy"})
        print(f"Bootstrap admin team updated: {BOOTSTRAP_ADMIN_EMAIL}")

ensure_bootstrap_admin()

# ── 認証ヘルパー ──────────────────────────────────────
def get_current_user():
    """セッションから現在のユーザー情報を返す。未ログインは None"""
    email = session.get("email")
    if not email:
        return None
    return {
        "email": email,
        "role": session.get("role", "viewer"),
        "display_name": session.get("display_name", email),
    }

def require_login(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            return jsonify({"error": "login required", "redirect": "/login"}), 401
        return f(*args, **kwargs)
    return wrapper

def require_editor(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"error": "login required", "redirect": "/login"}), 401
        if user["role"] != "editor":
            return jsonify({"error": "editor role required"}), 403
        return f(*args, **kwargs)
    return wrapper

# ── Firestore: 最終更新日時 ───────────────────────────
def get_last_refresh() -> dict | None:
    """Firestoreから最終BQ更新情報を取得"""
    if not _firestore_available:
        return None
    try:
        doc = db.collection(FIRESTORE_CONFIG_COLLECTION).document(REFRESH_DOC_ID).get()
        return doc.to_dict() if doc.exists else None
    except Exception as e:
        print(f"get_last_refresh error: {e}")
        return None

def set_last_refresh(iso_str: str, refreshed_by: str):
    """Firestoreに最終BQ更新情報を保存"""
    if not _firestore_available:
        return
    try:
        db.collection(FIRESTORE_CONFIG_COLLECTION).document(REFRESH_DOC_ID).set({
            "last_refresh": iso_str,
            "refreshed_by": refreshed_by,
        })
    except Exception as e:
        print(f"set_last_refresh error: {e}")

# ── BigQuery ベースクエリ ──────────────────────────────
BASE_QUERY = """
WITH raw01 AS (
  SELECT
    Booking_No_,
    ETD,
    Booking_Shipper,
    BKG_Shipper_code,
    POL_Sales,
    POL,
    CTR,
    CASE
      WHEN POL IN ("YOK","TYO","CHB","SMZ","NGO","THS","YKK","OSA","UKB") AND CTR = "KR" THEN "JPC_KR"
      WHEN POL NOT IN ("YOK","TYO","CHB","SMZ","NGO","THS","YKK","OSA","UKB") AND CTR = "KR" THEN "JPN_KR"
      ELSE CTR
    END AS AREA,
    POD,
    DLY,
    SUM(TEU) AS TEU
  FROM `updated_tables.score_dailybooking`
  WHERE ETD >= "2025-01-01"
  GROUP BY 1,2,3,4,5,6,7,8,9,10
  ORDER BY ETD
),
raw02 AS (
  SELECT DISTINCT BKG_No, BL_No
  FROM `updated_tables.lifting_detail`
  WHERE FORMAT_DATE('%Y-%m', DATE(Year, Month, 1)) >= '2025-01'
),
raw03 AS (
  SELECT DISTINCT
    FORMAT_DATE('%Y-%m', DATE(Year, MON, 1)) AS YearMonth,
    B_L_No_ AS BL_No,
    TEU,
    CM1
  FROM `updated_tables.lpa_revised_monthly`
  WHERE NOT (O_F IS NULL AND CM1 IS NULL)
)
SELECT
  FORMAT_DATE('%Y-%m', raw01.ETD) AS YearMonth,
  raw01.Booking_No_,
  raw02.BL_No,
  raw01.Booking_Shipper,
  raw01.BKG_Shipper_code,
  raw01.POL_Sales,
  raw01.POL,
  raw01.CTR,
  raw01.AREA,
  raw01.POD,
  raw01.DLY,
  raw01.ETD,
  SUM(raw01.TEU) AS TEU_score,
  SUM(raw03.TEU) AS TEU_lpa,
  SUM(raw03.CM1) AS CM1,
  SUM(CASE WHEN raw03.CM1 IS NOT NULL THEN raw03.TEU ELSE 0 END) AS TEU_with_cm1
FROM raw01
LEFT JOIN raw02 ON raw01.Booking_No_ = raw02.BKG_No
LEFT JOIN raw03 ON raw02.BL_No = raw03.BL_No
GROUP BY 1,2,3,4,5,6,7,8,9,10,11,12
ORDER BY BL_No
"""

def get_date_range_for_dashboard() -> tuple:
    """ダッシュボード用の日付範囲（過去6か月 〜 来月末）"""
    today = date.today()
    start = (today.replace(day=1) - timedelta(days=5*30)).replace(day=1)
    next_month_start = (today.replace(day=28) + timedelta(days=4)).replace(day=1)
    end = (next_month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    return start, end

# ── グローバル BQ キャッシュ ──────────────────────────
# データは起動時に1回ロード。以降はスケジューラーか手動更新のみ更新。
_bq_df: pd.DataFrame | None = None
_bq_loaded: bool = False
_bq_lock = threading.Lock()

_EMPTY_DF = pd.DataFrame(columns=[
    "ETD","YearMonth","Booking_No_","BL_No","Booking_Shipper","BKG_Shipper_code",
    "POL_Sales","POL","CTR","AREA","POD","DLY",
    "TEU_score","TEU_lpa","CM1","TEU_with_cm1","TEU",
    "week_year","week_no","week_key","month","ym","etd_ym","bq_ym"])

def _raw_fetch_bq() -> pd.DataFrame:
    """BigQuery から実際にデータを取得する（ロックなし）"""
    df = bq_client.query(BASE_QUERY).to_dataframe()
    df["ETD"] = pd.to_datetime(df["ETD"]).dt.date
    df["TEU_score"] = pd.to_numeric(df["TEU_score"], errors="coerce").fillna(0)
    df["TEU_lpa"]   = pd.to_numeric(df["TEU_lpa"],   errors="coerce")  # NaN を保持
    df["CM1"]       = pd.to_numeric(df["CM1"],        errors="coerce").fillna(0)
    # ① TEU_lpa があればそちらを優先、なければ TEU_score
    df["TEU"] = df["TEU_lpa"].fillna(df["TEU_score"])
    # ② CM1/TEU計算用: CM1が存在するレコードのみ表示用TEUをカウント
    df["TEU_with_cm1"] = df["TEU"].where(df["CM1"] != 0, 0)
    week_infos = df["ETD"].apply(get_week_info)
    df["week_year"]  = week_infos.apply(lambda x: x["year"])
    df["week_no"]    = week_infos.apply(lambda x: x["week"])
    df["week_key"]   = week_infos.apply(lambda x: x["week_key"])
    df["month"]      = week_infos.apply(lambda x: x["month"])
    df["ym"]         = week_infos.apply(lambda x: x["ym"])
    # 実カレンダー月 (ETD の年月) - 月間推移チャート用
    df["etd_ym"]     = df["ETD"].apply(lambda d: f"{d.year}-{d.month:02d}")
    # BQ YearMonth 列 (FORMAT_DATE('%Y-%m', ETD)) - 月間見込み・Top5用
    df["bq_ym"]      = df["YearMonth"].fillna(df["etd_ym"]).astype(str)
    return df

def do_refresh_bq(update_firestore: bool = True, refreshed_by: str = "system") -> bool:
    """BQデータをリフレッシュしてグローバルキャッシュに保存。
    update_firestore=True のときは Firestore の last_refresh も更新。"""
    global _bq_df, _bq_loaded
    with _bq_lock:
        try:
            df = _raw_fetch_bq()
            _bq_df = df
            _bq_loaded = True
            print(f"BQ data refreshed: {len(df)} rows")
            if update_firestore:
                now_jst = datetime.now(tz=JST).isoformat()
                set_last_refresh(now_jst, refreshed_by)
            return True
        except Exception as e:
            print(f"BigQuery refresh error: {e}")
            if not _bq_loaded:
                _bq_df = _EMPTY_DF.copy()
                _bq_loaded = True
            return False

# Summary から除外する荷主名
_EXCLUDE_SHIPPERS = {"RVTG", "THIS IS TEST BANK CODE"}

def get_bq_df() -> pd.DataFrame:
    """グローバルキャッシュを返す。未ロードなら起動時ロード（Firestore更新なし）"""
    if not _bq_loaded:
        do_refresh_bq(update_firestore=False, refreshed_by="startup")
    df = _bq_df if _bq_df is not None else _EMPTY_DF.copy()
    # テスト用・除外対象の荷主を除去
    if not df.empty and "Booking_Shipper" in df.columns:
        df = df[~df["Booking_Shipper"].isin(_EXCLUDE_SHIPPERS)]
    return df

# ローカルとの互換性のため fetch_bq_data も残す（引数は無視してキャッシュを返す）
def fetch_bq_data(start_date: date, end_date: date) -> pd.DataFrame:
    return get_bq_df()

# ── サブエリア定義 (PODベースの国内分割) ────────────────────
# key: サブエリア名, value: (親AREA, PODリスト, include=True/exclude=False)
SUB_AREA_DEFS = {
    "IN-West":  ("IN", ["NSA", "MUN"], True),     # NSA & MUN のみ
    "IN-East":  ("IN", ["NSA", "MUN"], False),     # NSA & MUN 以外
    "PKG&PEN":  ("MY", ["PKG", "PEN"], True),
    "PKW&PGU":  ("MY", ["PKW", "PGU"], True),
    "MIP":      ("PH", ["MIP"], True),
    "MNL":      ("PH", ["MNL"], True),
    "SGN":      ("VN", ["SGN"], True),
    "HPH":      ("VN", ["HPH"], True),
}

def _sub_area_parent(area: str) -> str | None:
    """サブエリアの親エリアを返す。サブエリアでなければ None。"""
    if area in SUB_AREA_DEFS:
        return SUB_AREA_DEFS[area][0]
    return None

def get_all_areas(df: pd.DataFrame) -> list:
    areas = sorted(df["AREA"].dropna().unique().tolist())
    # KR = JPC_KR + JPN_KR の合計として追加
    if "JPC_KR" in areas or "JPN_KR" in areas:
        areas = ["KR"] + areas
    # サブエリアは親エリアが存在すれば追加
    sub_areas = [sa for sa, (parent, _, _) in SUB_AREA_DEFS.items()
                 if parent in areas]
    all_available = set(areas) | set(sub_areas)
    # 表示順を固定
    preferred = [
        "KR", "JPC_KR", "JPN_KR", "AE", "NCN", "SCN",
        "HK", "ID",
        "IN", "IN-West", "IN-East",
        "PK",
        "MY", "PKG&PEN", "PKW&PGU",
        "PH", "MIP", "MNL",
        "SG", "TH",
        "VN", "SGN", "HPH",
        "MX", "US",
    ]
    ordered = [a for a in preferred if a in all_available]
    ordered += [a for a in areas if a not in preferred and a not in ordered]
    return ["ALL"] + ordered

# ── カレンダー月ユーティリティ ────────────────────────
def _cal_months(year: int, month: int, past: int, future: int) -> list:
    """実カレンダー月のリストを返す [(year, month), ...]"""
    res = []
    for i in range(past, 0, -1):
        m2, y2 = month - i, year
        while m2 <= 0:
            m2 += 12; y2 -= 1
        res.append((y2, m2))
    res.append((year, month))
    for i in range(1, future + 1):
        m2, y2 = month + i, year
        while m2 > 12:
            m2 -= 12; y2 += 1
        res.append((y2, m2))
    return res

# ── Firestore データアクセス ───────────────────────────
def load_prospects(week_keys: list, meeting_week: str = "") -> dict:
    """見込みデータを Firestore から取得。meeting_week で会議週スコープ。
    {week_key|area: {teu, cm1}}"""
    if not _firestore_available or not week_keys:
        return {}
    result = {}
    query = db.collection(FS_PROSPECTS)
    if meeting_week:
        query = query.where("meeting_week", "==", meeting_week)
    for i in range(0, len(week_keys), 30):
        chunk = week_keys[i:i+30]
        docs = query.where("week_key", "in", chunk).stream()
        for doc in docs:
            d = doc.to_dict()
            result[f"{d['week_key']}|{d['area']}"] = {"teu": d.get("teu"), "cm1": d.get("cm1")}
    return result

def load_monthly_prospects(yms: list, meeting_week: str = "") -> dict:
    """月間見込みを Firestore から取得。meeting_week で会議週スコープ。
    {ym|area: {teu, cm1_per_teu}}"""
    if not _firestore_available or not yms:
        return {}
    result = {}
    query = db.collection(FS_MONTHLY_PROSPECTS)
    if meeting_week:
        query = query.where("meeting_week", "==", meeting_week)
    for i in range(0, len(yms), 30):
        chunk = yms[i:i+30]
        docs = query.where("ym", "in", chunk).stream()
        for doc in docs:
            d = doc.to_dict()
            result[f"{d['ym']}|{d['area']}"] = {"teu": d.get("teu"), "cm1_per_teu": d.get("cm1_per_teu")}
    return result

def load_notes(week_key: str) -> dict:
    """メモを Firestore から取得。フォールバック: SQLite"""
    if _firestore_available:
        try:
            docs = db.collection(FS_NOTES).where("week_key", "==", week_key).stream()
            return {d.to_dict()["area"]: d.to_dict().get("note", "") for d in docs}
        except Exception as e:
            print(f"load_notes Firestore error: {e}")
    # fallback SQLite
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    rows = c.execute(
        "SELECT area, note FROM notes WHERE week_key = ?", (week_key,)
    ).fetchall()
    conn.close()
    return {area: note for area, note in rows}

def load_images(week_key: str, area: str) -> list:
    """画像メタデータを Firestore から取得。フォールバック: SQLite"""
    if _firestore_available:
        try:
            docs = (db.collection(FS_IMAGES)
                    .where("week_key", "==", week_key)
                    .where("area", "==", area)
                    .stream())
            return [{"id": d.id, "filename": d.to_dict()["filename"],
                     "caption": d.to_dict().get("caption", "")} for d in docs]
        except Exception as e:
            print(f"load_images Firestore error: {e}")
    # fallback SQLite
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    rows = c.execute(
        "SELECT id, filename, caption FROM images WHERE week_key=? AND area=?",
        (week_key, area)
    ).fetchall()
    conn.close()
    return [{"id": r[0], "filename": r[1], "caption": r[2]} for r in rows]

# ── API エンドポイント ─────────────────────────────────

@app.route("/")
def index():
    if not get_current_user():
        return redirect("/login")
    return send_from_directory("templates", "index.html")

@app.route("/login")
def login_page():
    return send_from_directory("templates", "login.html")

@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory("static", filename)

@app.route("/data/uploads/<path:filename>")
def uploaded_files(filename):
    """GCS（利用可能時）またはローカルから画像を配信"""
    if _gcs_available:
        blob = gcs_bucket.blob(f"meeting-uploads/{filename}")
        if blob.exists():
            content = blob.download_as_bytes()
            ct = blob.content_type or "image/png"
            return Response(content, content_type=ct)
        return "Not found", 404
    return send_from_directory(str(UPLOAD_DIR), filename)


@app.route("/api/auth/google", methods=["POST"])
def api_google_login():
    """Firebase IDトークンを検証してセッションを作成"""
    data = request.json or {}
    id_token = data.get("id_token", "")
    if not id_token:
        return jsonify({"error": "id_token required"}), 400

    try:
        decoded = fb_auth.verify_id_token(id_token)
    except Exception as e:
        return jsonify({"error": f"トークン検証失敗: {str(e)}"}), 401

    email = decoded.get("email", "")
    if not email:
        return jsonify({"error": "メールアドレスが取得できません"}), 400

    # Firestoreでユーザー確認
    fs_user = get_fs_user(email)
    if not fs_user:
        return jsonify({
            "error": f"このアカウント ({email}) はアクセス権限がありません。管理者に連絡してください。"
        }), 403

    # セッション作成
    display_name = decoded.get("name") or fs_user["display_name"]
    session["email"]        = email
    session["role"]         = fs_user["role"]
    session["display_name"] = display_name
    session.permanent       = False

    return jsonify({
        "email": email,
        "role":  fs_user["role"],
        "display_name": display_name,
    })

@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/auth/me")
def api_me():
    user = get_current_user()
    if not user:
        return jsonify({"logged_in": False}), 200
    return jsonify({"logged_in": True, **user})

@app.route("/api/auth/users", methods=["GET"])
@require_editor
def api_list_users():
    docs = db.collection(FIRESTORE_USERS_COLLECTION).stream()
    users = []
    for doc in docs:
        d = doc.to_dict()
        users.append({
            "email":        doc.id,
            "role":         d.get("role", "viewer"),
            "display_name": d.get("display_name", doc.id),
        })
    users.sort(key=lambda u: (u["role"] != "editor", u["email"]))
    return jsonify(users)

@app.route("/api/auth/users", methods=["POST"])
@require_editor
def api_add_user():
    data = request.json or {}
    email        = data.get("email", "").strip().lower()
    role         = data.get("role", "viewer")
    display_name = data.get("display_name", email)
    if not email:
        return jsonify({"error": "email required"}), 400
    if role not in ("editor", "viewer"):
        return jsonify({"error": "role must be editor or viewer"}), 400
    me = get_current_user()
    db.collection(FIRESTORE_USERS_COLLECTION).document(email).set({
        "role":         role,
        "display_name": display_name,
        "created_by":   me["email"] if me else "unknown",
    })
    return jsonify({"ok": True})

@app.route("/api/auth/users/<path:email>", methods=["DELETE"])
@require_editor
def api_delete_user(email):
    me = get_current_user()
    if me and me["email"] == email:
        return jsonify({"error": "自分自身は削除できません"}), 400
    db.collection(FIRESTORE_USERS_COLLECTION).document(email).delete()
    return jsonify({"ok": True})

@app.route("/api/auth/users/<path:email>/role", methods=["POST"])
@require_editor
def api_change_role(email):
    data = request.json or {}
    new_role = data.get("role", "viewer")
    if new_role not in ("editor", "viewer"):
        return jsonify({"error": "invalid role"}), 400
    db.collection(FIRESTORE_USERS_COLLECTION).document(email).update({"role": new_role})
    return jsonify({"ok": True})


@app.route("/api/areas")
@require_login
def api_areas():
    start, end = get_date_range_for_dashboard()
    df = fetch_bq_data(start, end)
    return jsonify({"areas": get_all_areas(df)})


def _week_key_to_date(week_key: str) -> date | None:
    """week_key (例: '2026-W16') → その週の開始日 (date) を返す"""
    import re
    m = re.match(r"(\d{4})-W(\d+)", week_key or "")
    if not m:
        return None
    yr, wk = int(m.group(1)), int(m.group(2))
    ws, _ = get_week_date_range(yr, wk)
    return ws


def _ref_date_from_meeting_week(meeting_week: str, today: date) -> date:
    """meeting_week から基準日を算出。チャート範囲のシフトに使用。
    meeting_week が空や無効ならば today を返す。"""
    if not meeting_week:
        return today
    d = _week_key_to_date(meeting_week)
    return d if d else today


def build_summary_for_area(area: str, df: pd.DataFrame, today: date,
                           meeting_week: str = "") -> dict:
    """エリア別サマリーを構築（api_summary / archive 共通ロジック）
    meeting_week: 会議週 (例: "2026-W14")。見込みデータのスコープに使用。
    チャート範囲は meeting_week の月を基準にシフトする。"""
    # チャート基準日: 選択週の月に合わせる
    ref = _ref_date_from_meeting_week(meeting_week, today)
    # エリアフィルタ (KR は JPC_KR + JPN_KR の合計、サブエリアはPODで絞込)
    df_area = _filter_area(df, area)

    empty = {"monthly": [], "weekly": [], "top_shippers": [],
             "shipper_count": 0, "prospects": {}}
    if df_area.empty:
        return empty

    # ── 月次集計 (BQ YearMonth列ベース: TEU/CM1合計) ──────
    # 会議日(火曜)基準で月間推移の範囲を切替 (常に6か月表示)
    meeting_tue_day = (ref + timedelta(days=2)).day
    if meeting_tue_day <= 14:
        # ≤14日: 過去5か月+当月 (来月なし)
        months_list = _cal_months(ref.year, ref.month, 5, 0)
    else:
        # ≥15日: 過去4か月+当月+来月
        months_list = _cal_months(ref.year, ref.month, 4, 1)
    yms_monthly = [f"{y}-{m:02d}" for y, m in months_list]
    monthly_prospects_db = load_monthly_prospects(yms_monthly, meeting_week)

    # ref の月を基準に is_current / is_future を判定
    ref_ym = (ref.year, ref.month)
    monthly = []
    for y, m in months_list:
        ym_str = f"{y}-{m:02d}"
        sub = df_area[df_area["bq_ym"] == ym_str]
        teu = float(sub["TEU"].sum())
        cm1 = float(sub["CM1"].sum())
        teu_wcm1 = float(sub["TEU_with_cm1"].sum())
        is_fut  = (y, m) > ref_ym
        is_curr = (y, m) == ref_ym
        mp = monthly_prospects_db.get(f"{ym_str}|{area}", {})
        monthly.append({
            "ym": ym_str,
            "label": month_label(y, m),
            "year": y, "month": m,
            "TEU": round(teu),
            "CM1": round(cm1),
            "CM1_per_TEU": round(cm1 / teu_wcm1) if teu_wcm1 > 0 else 0,
            "shipper_count": int(sub["Booking_Shipper"].nunique()),
            "is_future": is_fut,
            "is_current": is_curr,
            "m_prospect_teu": mp.get("teu"),
            "m_prospect_cm1": mp.get("cm1_per_teu"),
        })

    # ── 週次集計 (会議日基準で2か月分) ────
    weeks_3m = get_2month_weeks(ref)
    all_week_keys = [w["week_key"] for w in weeks_3m]
    prospects_db = load_prospects(all_week_keys, meeting_week)

    weekly = []
    for w in weeks_3m:
        wk = w["week"]
        wy = w["year"]
        wkey = w["week_key"]
        ws = w["week_start"]
        we = w["week_end"]

        sub = df_area[df_area["week_key"] == wkey]
        teu = float(sub["TEU"].sum())
        cm1 = float(sub["CM1"].sum())
        teu_wcm1 = float(sub["TEU_with_cm1"].sum())
        w_shipper_count = int(sub["Booking_Shipper"].nunique()) if not sub.empty else 0

        p = prospects_db.get(f"{wkey}|{area}", {})
        p_teu = p.get("teu")
        p_cm1 = p.get("cm1")

        ws_date = date.fromisoformat(ws) if ws else None
        is_future = ws_date > ref if ws_date else False
        is_current = (ws_date <= ref <= date.fromisoformat(we)) if ws_date and we else False

        weekly.append({
            "week_key": wkey,
            "week": wk,
            "year": wy,
            "month": w["month"],
            "month_label": w["month_label"],
            "ym": w["ym"],
            "week_label": f"W{wk}",
            "week_start": ws,
            "week_end": we,
            "TEU": round(teu),
            "CM1": round(cm1),
            "CM1_per_TEU": round(cm1 / teu_wcm1) if teu_wcm1 > 0 else 0,
            "shipper_count": w_shipper_count,
            "prospect_TEU": p_teu,
            "prospect_CM1": p_cm1,
            "is_future": is_future,
            "is_current": is_current,
        })

    # ── Top5 Shipper (BQ YearMonth列ベース) ─────────────
    months_6m = _cal_months(ref.year, ref.month, 5, 0)
    ym_6m = [f"{y}-{m:02d}" for y, m in months_6m]
    df_6m = df_area[df_area["bq_ym"].isin(ym_6m)]

    shipper_total = (
        df_6m.groupby("Booking_Shipper")["TEU"].sum()
        .sort_values(ascending=False)
        .head(5)
    )
    top5_shippers = shipper_total.index.tolist()

    months_4 = _cal_months(ref.year, ref.month, 2, 1)  # 前々月〜翌月 (4か月分)
    prev2_y, prev2_m = months_4[0]
    prev_y, prev_m = months_4[1]
    curr_y, curr_m = months_4[2]
    next_y, next_m = months_4[3]

    def shipper_stats(shipper, sy, sm):
        ym_str = f"{sy}-{sm:02d}"
        sub = df_area[(df_area["Booking_Shipper"] == shipper) & (df_area["bq_ym"] == ym_str)]
        teu = float(sub["TEU"].sum())
        cm1 = float(sub["CM1"].sum())
        teu_wcm1 = float(sub["TEU_with_cm1"].sum())
        return {"TEU": round(teu), "CM1_per_TEU": round(cm1 / teu_wcm1) if teu_wcm1 > 0 else 0}

    # 3M平均TEU: 会議日ルールに基づく3か月を使用
    meeting_tue_date_top5 = ref + timedelta(days=2)
    if meeting_tue_date_top5.day <= 14:
        # 当月を含まない過去3か月 (例: W14=4/7 → 1,2,3月)
        months_3m_avg = _cal_months(ref.year, ref.month, 3, 0)[:-1]
    else:
        # 当月を含む過去3か月 (例: W16=4/21 → 2,3,4月)
        months_3m_avg = _cal_months(ref.year, ref.month, 2, 0)
    ym_3m_avg = [f"{y}-{m:02d}" for y, m in months_3m_avg]
    df_3m_avg = df_area[df_area["bq_ym"].isin(ym_3m_avg)]

    top_shippers = []
    for s in top5_shippers:
        prev2_s = shipper_stats(s, prev2_y, prev2_m)
        prev_s = shipper_stats(s, prev_y, prev_m)
        curr_s = shipper_stats(s, curr_y, curr_m)
        next_s = shipper_stats(s, next_y, next_m)
        # 3M平均TEU
        s_3m = df_3m_avg[df_3m_avg["Booking_Shipper"] == s]
        avg_3m_teu = round(float(s_3m["TEU"].sum()) / 3) if len(months_3m_avg) == 3 else 0
        top_shippers.append({
            "shipper": s,
            "total_6m_TEU": round(float(shipper_total[s])),
            "prev2": prev2_s, "prev": prev_s, "curr": curr_s, "next": next_s,
            "gap_teu": curr_s["TEU"] - prev_s["TEU"],
            "avg_3m_teu": avg_3m_teu,
            "avg_3m_months": [f"{y}/{m:02d}" for y, m in months_3m_avg],
        })

    curr_ym = f"{ref.year}-{ref.month:02d}"
    df_curr = df_area[df_area["bq_ym"] == curr_ym]
    shipper_count = int(df_curr["Booking_Shipper"].nunique())

    # 会議日（火曜日）の日付を算出
    meeting_tue_date = ref + timedelta(days=2)

    return {
        "monthly": monthly,
        "weekly": weekly,
        "top_shippers": top_shippers,
        "shipper_count": shipper_count,
        "prospects": {k.split("|")[0]: v for k, v in prospects_db.items() if k.endswith(f"|{area}")},
        "meeting_day": meeting_tue_date.day,
    }


@app.route("/api/download-data")
@require_login
def api_download_data():
    """ETDベースのローデータをExcelでダウンロード"""
    import io
    from openpyxl.utils import get_column_letter

    etd_from = request.args.get("from", "")
    etd_to = request.args.get("to", "")
    areas_param = request.args.get("areas", "")  # カンマ区切り。空=全エリア

    if not etd_from or not etd_to:
        return jsonify({"error": "from, to パラメータが必要です"}), 400

    df = get_bq_df()
    if df.empty:
        return jsonify({"error": "データが読み込まれていません"}), 500

    # ETD でフィルタ
    mask = (df["ETD"] >= date.fromisoformat(etd_from)) & (df["ETD"] <= date.fromisoformat(etd_to))
    filtered = df[mask].copy()

    # Area フィルタ
    if areas_param:
        area_list = [a.strip() for a in areas_param.split(",") if a.strip()]
        # KR = JPC_KR + JPN_KR
        expanded = set()
        for a in area_list:
            if a == "KR":
                expanded.add("JPC_KR")
                expanded.add("JPN_KR")
            elif a in SUB_AREA_DEFS:
                # サブエリアは親エリアでフィルタ後にPODで絞る
                parent, pods, _ = SUB_AREA_DEFS[a]
                expanded.add(parent + "::sub::" + a)
            else:
                expanded.add(a)
        # 通常エリア
        normal_areas = {a for a in expanded if "::sub::" not in a}
        sub_requests = {a.split("::sub::")[1]: a.split("::sub::")[0] for a in expanded if "::sub::" in a}

        parts = []
        if normal_areas:
            parts.append(filtered[filtered["AREA"].isin(normal_areas)])
        for sub_name, parent in sub_requests.items():
            _, pods, _ = SUB_AREA_DEFS[sub_name]
            sub_df = filtered[(filtered["AREA"] == parent) & (filtered["POD"].isin(pods))]
            parts.append(sub_df)
        filtered = pd.concat(parts, ignore_index=True) if parts else filtered.iloc[0:0]

    if filtered.empty:
        return jsonify({"error": "該当データがありません"}), 404

    # 出力カラム選定・整形
    # bq_ym = ダッシュボード月間推移と同じカレンダー月 (ETD実月)
    # ym    = 会社週カレンダー月 (週間推移チャートで使用)
    filtered = filtered.copy()
    filtered.rename(columns={"bq_ym": "Month(Calendar)", "ym": "ETD(445)"}, inplace=True)
    out_cols = [
        "ETD", "AREA", "CTR", "Booking_No_", "BL_No",
        "Booking_Shipper", "BKG_Shipper_code", "POL_Sales",
        "POL", "POD", "DLY",
        "week_key", "Month(Calendar)", "ETD(445)",
        "TEU", "CM1",
    ]
    existing = [c for c in out_cols if c in filtered.columns]
    out_df = filtered[existing].copy()
    out_df = out_df.sort_values(["ETD", "AREA", "POL", "POD"]).reset_index(drop=True)

    # Excel 書き出し
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="RawData")
        ws = writer.sheets["RawData"]
        for i, col in enumerate(out_df.columns, 1):
            max_len = max(len(str(col)), out_df[col].astype(str).str.len().max() if len(out_df) > 0 else 0)
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 30)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
    buf.seek(0)

    fname = f"meeting_data_{etd_from}_{etd_to}.xlsx"
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.route("/api/summary")
@require_login
def api_summary():
    """エリア別ダッシュボードデータを返す ?area=KR&meeting_week=2026-W14"""
    today = date.today()
    df = get_bq_df()
    area = request.args.get("area", "ALL")
    meeting_week = request.args.get("meeting_week", "")
    result = build_summary_for_area(area, df, today, meeting_week)
    # 自動合算エリアフラグを付与
    result["is_auto_sum_area"] = area in AREA_AUTO_SUM
    # サブエリア統合表示: 子エリアの weekly/monthly データも返す
    if area in AREA_AUTO_SUM:
        children = AREA_AUTO_SUM[area]
        sub_data = {}
        for child in children:
            child_result = build_summary_for_area(child, df, today, meeting_week)
            sub_data[child] = {
                "weekly": child_result.get("weekly", []),
                "monthly": child_result.get("monthly", []),
            }
        result["sub_areas"] = sub_data
        result["sub_area_names"] = children
    return jsonify(result)


# ── 予測 API ─────────────────────────────────────────────
# 4年分の軽量BQクエリ (予測専用: TEU_lpa + CM1 のみ)
PREDICT_QUERY = """
WITH raw01 AS (
  SELECT Booking_No_, ETD, BKG_Date, POL, POD,
    CASE
      WHEN POL IN ("YOK","TYO","CHB","SMZ","NGO","THS","YKK","OSA","UKB") AND CTR = "KR" THEN "JPC_KR"
      WHEN POL NOT IN ("YOK","TYO","CHB","SMZ","NGO","THS","YKK","OSA","UKB") AND CTR = "KR" THEN "JPN_KR"
      ELSE CTR
    END AS AREA,
    SUM(TEU) AS TEU
  FROM `updated_tables.score_dailybooking`
  WHERE ETD >= DATE_TRUNC(DATE_SUB(CURRENT_DATE(), INTERVAL 4 YEAR), YEAR)
  GROUP BY 1,2,3,4,5,6
),
raw02 AS (
  SELECT DISTINCT BKG_No, BL_No
  FROM `updated_tables.lifting_detail`
  WHERE FORMAT_DATE('%Y-%m', DATE(Year, Month, 1)) >= FORMAT_DATE('%Y-%m', DATE_TRUNC(DATE_SUB(CURRENT_DATE(), INTERVAL 4 YEAR), YEAR))
),
raw03 AS (
  SELECT DISTINCT FORMAT_DATE('%Y-%m', DATE(Year, MON, 1)) AS YearMonth,
    B_L_No_ AS BL_No, TEU, CM1
  FROM `updated_tables.lpa_revised_monthly`
  WHERE NOT (O_F IS NULL AND CM1 IS NULL)
)
SELECT FORMAT_DATE('%Y-%m', raw01.ETD) AS YearMonth,
  raw01.AREA, raw01.POD, raw01.ETD, raw01.BKG_Date,
  SUM(raw03.TEU) AS TEU_lpa,
  SUM(raw03.CM1) AS CM1
FROM raw01
LEFT JOIN raw02 ON raw01.Booking_No_ = raw02.BKG_No
LEFT JOIN raw03 ON raw02.BL_No = raw03.BL_No
GROUP BY 1,2,3,4,5
ORDER BY ETD
"""

# 予測用データキャッシュ (ボタン押下時にBQ問い合わせ → 1時間キャッシュ)
_predict_df: pd.DataFrame | None = None
_predict_loaded_at: float = 0
_predict_lock = threading.Lock()
_PREDICT_CACHE_SEC = 3600  # 1時間

def _fetch_predict_df() -> pd.DataFrame:
    """予測用4年分データをBQから取得 (キャッシュ付き)"""
    global _predict_df, _predict_loaded_at
    import time
    now = time.time()
    with _predict_lock:
        if _predict_df is not None and (now - _predict_loaded_at) < _PREDICT_CACHE_SEC:
            return _predict_df
    df = bq_client.query(PREDICT_QUERY).to_dataframe()
    df["ETD"] = pd.to_datetime(df["ETD"])
    df["BKG_Date"] = pd.to_datetime(df["BKG_Date"], errors="coerce")
    df["TEU_lpa"] = pd.to_numeric(df["TEU_lpa"], errors="coerce")
    df["CM1"] = pd.to_numeric(df["CM1"], errors="coerce").fillna(0)
    # CM1/TEU計算用: CM1が存在するレコードのみTEU_lpaをカウント
    df["TEU_with_cm1"] = df["TEU_lpa"].where(df["CM1"] != 0, 0)
    print(f"Predict BQ: {len(df):,} rows, ETD {df['ETD'].min().date()} ~ {df['ETD'].max().date()}")
    with _predict_lock:
        _predict_df = df
        _predict_loaded_at = now
    return df


def _build_prediction(area: str, today: date,
                      meeting_week: str = "") -> dict:
    """4年分の TEU_lpa データで季節性分析 → 将来の TEU / CM1 を予測。
    月間: 月別季節性指数 × 直近6か月ベースライン (TEU_lpa のみ)
    CM1/TEU: 直近6か月の線形回帰でトレンド延長
    週間: 季節別(冬/夏/通常)の月内均等セグメント配分 (4年分フル活用)"""
    ref = _ref_date_from_meeting_week(meeting_week, today)
    df = _fetch_predict_df()

    # エリアフィルタ (予測クエリにPOD列なし → サブエリアは親エリアにフォールバック)
    df_area = _filter_area_no_pod(df, area)

    if df_area.empty:
        return {"monthly": {}, "weekly": {}}

    # ── 月別集計 (TEU_lpa のみ) ───────────────────────
    df_area["_ym"] = df_area["YearMonth"]
    df_area["_cal_month"] = df_area["ETD"].dt.month

    monthly_agg = df_area.groupby("_ym").agg(
        TEU_lpa=("TEU_lpa", "sum"), CM1=("CM1", "sum"),
        TEU_with_cm1=("TEU_with_cm1", "sum")
    ).reset_index()
    monthly_agg["CM1_per_TEU"] = monthly_agg.apply(
        lambda r: round(r["CM1"] / r["TEU_with_cm1"]) if r["TEU_with_cm1"] > 0 else 0, axis=1)
    monthly_agg["month_num"] = monthly_agg["_ym"].apply(lambda s: int(s.split("-")[1]))

    # 不完全月を除外 (当月以降 + データ欠損月)
    ref_ym = f"{ref.year}-{ref.month:02d}"
    complete = monthly_agg[monthly_agg["_ym"] < ref_ym].copy()

    # 外れ値除外: 各月の中央値から±2σ以上離れた月を除外
    if len(complete) > 6:
        for m_num in range(1, 13):
            sub = complete[complete["month_num"] == m_num]["TEU_lpa"]
            if len(sub) >= 2:
                median_val = sub.median()
                std_val = sub.std()
                if std_val > 0:
                    complete = complete[~(
                        (complete["month_num"] == m_num) &
                        ((complete["TEU_lpa"] - median_val).abs() > 2 * std_val)
                    )]

    # 季節性指数: 各月(1-12)の平均 TEU_lpa / 全体平均
    overall_avg = complete["TEU_lpa"].mean() if len(complete) > 0 else 1
    seasonal_idx = {}
    for m_num in range(1, 13):
        sub = complete[complete["month_num"] == m_num]
        if len(sub) > 0:
            seasonal_idx[m_num] = float(sub["TEU_lpa"].mean()) / overall_avg if overall_avg > 0 else 1.0
        else:
            seasonal_idx[m_num] = 1.0

    # 直近6か月の TEU_lpa (季節性除去してベースライン算出)
    # CM1/TEU はトレンド（下落・上昇）を捉えるため加重平均 + 線形回帰
    recent_months = _cal_months(ref.year, ref.month, 6, 0)[:-1]  # 当月除く過去6か月
    recent_teus = []
    recent_cm1s = []  # [(month_offset, cm1_per_teu)]
    for idx, (y, m) in enumerate(recent_months):
        ym_str = f"{y}-{m:02d}"
        sub = monthly_agg[monthly_agg["_ym"] == ym_str]
        if len(sub) > 0 and float(sub["TEU_lpa"].iloc[0]) > 0:
            raw_teu = float(sub["TEU_lpa"].iloc[0])
            raw_cm1t = float(sub["CM1_per_TEU"].iloc[0])
            si = seasonal_idx.get(m, 1.0)
            recent_teus.append(raw_teu / si if si > 0 else raw_teu)
            recent_cm1s.append((idx, raw_cm1t))

    base_teu = sum(recent_teus) / len(recent_teus) if recent_teus else overall_avg

    # CM1/TEU: 線形回帰で直近トレンドを延長予測
    if len(recent_cm1s) >= 3:
        # 最小二乗法 (y = a*x + b)
        n = len(recent_cm1s)
        xs = [c[0] for c in recent_cm1s]
        ys = [c[1] for c in recent_cm1s]
        x_mean = sum(xs) / n
        y_mean = sum(ys) / n
        ss_xy = sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys))
        ss_xx = sum((x - x_mean) ** 2 for x in xs)
        slope = ss_xy / ss_xx if ss_xx > 0 else 0
        intercept = y_mean - slope * x_mean
        # 予測月のオフセット (直近月の次 = len(recent_months))
        base_offset = len(recent_months)
    else:
        slope = 0
        intercept = sum(c[1] for c in recent_cm1s) / len(recent_cm1s) if recent_cm1s else 0
        base_offset = 0

    # ── 当月の実績 & BKG_Dateベースの残Booking推定 ────────
    ref_ym_str = f"{ref.year}-{ref.month:02d}"
    import calendar as _cal_mod
    days_in_ref_month = _cal_mod.monthrange(ref.year, ref.month)[1]
    elapsed_days = ref.day

    # 当月の全実績TEU (全ETD含む = 既にBooking済みの全量)
    curr_actual_all = monthly_agg[monthly_agg["_ym"] == ref_ym_str]
    curr_actual_teu_all = float(curr_actual_all["TEU_lpa"].iloc[0]) if len(curr_actual_all) > 0 else 0

    # ── BKG_Date分析: 週ポジション別の残りBooking率を学習 ──
    # エリアごと × 週ポジション(0=第1週,1=第2週...)ごとに
    # 「day X時点でまだBookingされていなかった割合」をデータから算出
    week_remaining_rates = {}  # {week_pos: remaining_rate}
    monthly_remaining_rate = 0.0  # 月全体のフォールバック
    try:
        hist = df_area[
            (df_area["ETD"] < pd.Timestamp(ref.year, ref.month, 1)) &
            (df_area["BKG_Date"].notna()) &
            (df_area["TEU_lpa"].notna())
        ].copy()
        if len(hist) > 100:
            hist["_etd_ym"] = hist["ETD"].dt.to_period("M")
            hist["_etd_month_start"] = hist["_etd_ym"].dt.start_time
            hist["_cutoff"] = hist["_etd_month_start"] + pd.Timedelta(days=elapsed_days - 1)
            hist["_was_visible"] = hist["BKG_Date"] <= hist["_cutoff"]

            # ETDの日付から週ポジション(0-based)を計算
            hist["_etd_day"] = hist["ETD"].dt.day
            hist["_days_in_m"] = hist["ETD"].apply(
                lambda d: _cal_mod.monthrange(d.year, d.month)[1])
            hist["_week_pos"] = ((hist["_etd_day"] - 1) * 4 // hist["_days_in_m"]).clip(0, 3)
            # 4分割: 0=day1-7, 1=day8-14, 2=day15-22, 3=day23-末

            # 週ポジションごとに残り率を集計
            for wpos in range(4):
                pos_rates = []
                for _, grp in hist[hist["_week_pos"] == wpos].groupby("_etd_ym"):
                    total = grp["TEU_lpa"].sum()
                    if total <= 0:
                        continue
                    visible = grp.loc[grp["_was_visible"], "TEU_lpa"].sum()
                    pos_rates.append(1.0 - visible / total)
                if pos_rates:
                    pos_rates.sort()
                    mid = len(pos_rates) // 2
                    median_rate = pos_rates[mid] if len(pos_rates) % 2 else (pos_rates[mid-1] + pos_rates[mid]) / 2
                    week_remaining_rates[wpos] = max(0, median_rate)

            # 月全体の残り率 (月間予測補正用)
            all_rates = []
            for _, grp in hist.groupby("_etd_ym"):
                total = grp["TEU_lpa"].sum()
                if total <= 0:
                    continue
                visible = grp.loc[grp["_was_visible"], "TEU_lpa"].sum()
                all_rates.append(1.0 - visible / total)
            if all_rates:
                all_rates.sort()
                mid = len(all_rates) // 2
                monthly_remaining_rate = all_rates[mid] if len(all_rates) % 2 else (all_rates[mid-1] + all_rates[mid]) / 2

            print(f"Predict [{area}] at day {elapsed_days}: "
                  f"monthly_remaining={monthly_remaining_rate:.1%}, "
                  f"week_pos_rates={{{', '.join(f'W{k+1}:{v:.1%}' for k,v in sorted(week_remaining_rates.items()))}}}, "
                  f"from {len(all_rates)} months history")
    except Exception as e:
        print(f"BKG_Date analysis error: {e}")
        # フォールバック
        for wpos in range(4):
            day_start = wpos * (days_in_ref_month // 4) + 1
            days_until = max(0, day_start - elapsed_days)
            week_remaining_rates[wpos] = max(0, days_until / days_in_ref_month) ** 2 * 0.15

    # ── ダッシュボード実績 (TEU) を月間フロアにも使用 ────────────
    try:
        actual_df = get_bq_df()
        actual_df_area = _filter_area(actual_df, area) if area != "ALL" else actual_df
    except Exception:
        actual_df_area = pd.DataFrame()

    # ダッシュボード上の当月実績TEU (TEU_lpaより大きい場合がある)
    dash_actual_monthly = {}
    if not actual_df_area.empty and "ym" in actual_df_area.columns:
        dash_monthly = actual_df_area.groupby("ym")["TEU"].sum()
        dash_actual_monthly = {k: float(v) for k, v in dash_monthly.items()}

    # ── 月間予測 ────────────────────────────────────────
    predict_months = _cal_months(ref.year, ref.month, 0, 1)
    monthly_pred = {}
    _cm1_by_month_idx = {}
    for i, (y, m) in enumerate(predict_months):
        ym_str = f"{y}-{m:02d}"
        si = seasonal_idx.get(m, 1.0)
        pred_teu_raw = base_teu * si  # 基本予測: 季節性指数ベース

        # 補正: 当月は実績ベースで着地予測
        if ym_str == ref_ym_str and curr_actual_teu_all > 0:
            # 実績 + 残りBooking推定（月全体の残り率）
            remaining_teu = curr_actual_teu_all * monthly_remaining_rate
            actual_based = curr_actual_teu_all + remaining_teu
            # 上限: 過去同月の最大値×1.2
            same_month_max = float(complete[complete["month_num"] == m]["TEU_lpa"].max()) if len(complete[complete["month_num"] == m]) > 0 else pred_teu_raw * 1.5
            cap = same_month_max * 1.2
            pred_teu_raw = min(max(pred_teu_raw, actual_based), cap)
            # 絶対下限: 当月実績を下回らない
            pred_teu_raw = max(pred_teu_raw, curr_actual_teu_all)

        # CM1/TEU: トレンド延長
        pred_cm1_raw = slope * (base_offset + i) + intercept
        if pred_cm1_raw < 0:
            pred_cm1_raw = intercept
        # ダッシュボード実績でも絶対下限を保証
        dash_actual = dash_actual_monthly.get(ym_str, 0)
        pred_teu_raw = max(pred_teu_raw, dash_actual)

        monthly_pred[ym_str] = {
            "teu": round(pred_teu_raw),
            "cm1_per_teu": round(pred_cm1_raw),
        }
        _cm1_by_month_idx[ym_str] = pred_cm1_raw

    # ── 週間予測 (季節別・均等セグメント配分、4年分フル活用) ────
    # 会社カレンダー不要: 各月の日数をN等分してTEU_lpaの配分比率を学習
    # 季節区分: 冬(12,1,2,3)=船遅延/月初膨張, 夏(6,7,8,9)=台風遅延, 通常(4,5,10,11)
    import calendar as _cal

    def _predict_season(cal_month: int) -> str:
        if cal_month in (12, 1, 2, 3):
            return "winter"
        elif cal_month in (6, 7, 8, 9):
            return "summer"
        return "normal"

    def _assign_segment(day: int, days_in_month: int, n_segments: int) -> int:
        seg_size = days_in_month / n_segments
        return min(int((day - 1) / seg_size), n_segments - 1)

    # 当月以前のデータで学習
    df_train = df_area[df_area["ETD"] < pd.Timestamp(ref.year, ref.month, 1)].copy()
    df_train["_day"] = df_train["ETD"].dt.day
    df_train["_year"] = df_train["ETD"].dt.year
    df_train["_cal_month"] = df_train["ETD"].dt.month
    df_train["_days_in_month"] = df_train.apply(
        lambda r: _cal.monthrange(int(r["_year"]), int(r["_cal_month"]))[1], axis=1)

    # 4分割・5分割の両方でセグメント配分を学習
    seg_ratios = {}     # {(n_seg, season): {pos: ratio}}
    seg_ratios_all = {} # {n_seg: {pos: ratio}}
    for n_seg in [4, 5]:
        df_s = df_train.copy()
        df_s["_seg"] = df_s.apply(
            lambda r: _assign_segment(int(r["_day"]), int(r["_days_in_month"]), n_seg), axis=1)
        seg_agg = df_s.groupby(["_ym", "_seg", "_cal_month"]).agg(
            TEU_lpa=("TEU_lpa", "sum")).reset_index()
        month_tots = seg_agg.groupby("_ym")["TEU_lpa"].sum()
        seg_agg["month_total"] = seg_agg["_ym"].map(month_tots)
        seg_agg = seg_agg[seg_agg["month_total"] > 0]
        seg_agg["ratio"] = seg_agg["TEU_lpa"] / seg_agg["month_total"]
        seg_agg["season"] = seg_agg["_cal_month"].apply(_predict_season)

        # 全体 (非季節)
        seg_ratios_all[n_seg] = seg_agg.groupby("_seg")["ratio"].mean().to_dict()
        # 季節別
        for sn in ["winter", "summer", "normal"]:
            sub = seg_agg[seg_agg["season"] == sn]
            if len(sub) > 0:
                seg_ratios[(n_seg, sn)] = sub.groupby("_seg")["ratio"].mean().to_dict()

    # 週間予測: 各将来月の週に配分
    weeks_3m = get_3month_weeks(ref)

    # 週別実績を取得 (当月の週で実績がある分を把握)
    weekly_actuals = {}  # {week_key: actual_teu}
    if not actual_df_area.empty and "week_key" in actual_df_area.columns:
        wa = actual_df_area.groupby("week_key")["TEU"].sum()
        weekly_actuals = {k: float(v) for k, v in wa.items()}

    weekly_pred = {}
    for ym_str in set(w["ym"] for w in weeks_3m):
        m_pred = monthly_pred.get(ym_str)
        if not m_pred:
            continue
        month_weeks = [w for w in weeks_3m if w["ym"] == ym_str]
        n_weeks = len(month_weeks)
        cal_m = int(ym_str.split("-")[1])
        season = _predict_season(cal_m)

        # 季節別比率を優先、なければ全体比率
        ratios = seg_ratios.get((n_weeks, season)) or seg_ratios_all.get(n_weeks) or {}
        if not ratios:
            ratios = {i: 1.0 / n_weeks for i in range(n_weeks)}

        raw_shares = [ratios.get(i, 1.0 / n_weeks) for i in range(n_weeks)]
        total_r = sum(raw_shares)

        # ── 補正: 実績がある週を考慮して残りを再配分 ──
        # 各週の基本配分を計算
        base_alloc = []
        for pos_idx, w in enumerate(month_weeks):
            share = raw_shares[pos_idx] / total_r if total_r > 0 else 1.0 / n_weeks
            base_alloc.append({
                "w": w,
                "share": share,
                "base_teu": int(m_pred["teu"] * share + 0.5),
            })

        # 実績が既にある週を特定し、残りTEUを将来週に再配分
        # Note: 先行Bookingにより、未来週にも既に実績データがある場合がある
        actual_sum = 0  # 実績がある週の合計
        future_shares_sum = 0  # 将来週のシェア合計
        for ba in base_alloc:
            wk = ba["w"]["week_key"]
            actual_teu = weekly_actuals.get(wk, 0)
            if actual_teu > 0:
                # 実績データがある週（week_startが未来でもBooking済み）
                ba["has_actual"] = True
                ba["actual_teu"] = actual_teu
                actual_sum += actual_teu
            else:
                ba["has_actual"] = False
                future_shares_sum += ba["share"]

        # 月間予測から実績分を差し引いた残りを将来週に配分
        remaining_teu = max(0, m_pred["teu"] - actual_sum)

        for ba in base_alloc:
            wk = ba["w"]["week_key"]
            if ba["has_actual"]:
                # 実績がある週: 実績 + 週ポジション別の残りBooking率を加味
                ws = ba["w"].get("week_start")
                ws_date = date.fromisoformat(ws) if ws else None
                if ws_date and ws_date > ref:
                    # 未来週: BKG_Dateデータに基づく週ポジション別の残り率
                    ws_dim = _cal_mod.monthrange(ws_date.year, ws_date.month)[1]
                    week_mid_day = ws_date.day + 3
                    wpos = min(3, (week_mid_day - 1) * 4 // ws_dim)
                    week_rate = week_remaining_rates.get(wpos, 0.0)
                    # 最低でも10%増し (まだBookingは入ってくる)
                    effective_rate = max(week_rate, 0.10)
                    actual_plus = ba["actual_teu"] * (1 + effective_rate)
                    pred_teu = int(max(actual_plus, ba["base_teu"]) + 0.5)
                else:
                    # 当週: 実績×1.05 (今週中にまだBookingは入る) vs 基本予測の大きい方
                    actual_plus_curr = int(ba["actual_teu"] * 1.05 + 0.5)
                    pred_teu = max(actual_plus_curr, ba["base_teu"])
            elif future_shares_sum > 0 and actual_sum > 0:
                # 実績ゼロの将来週: 残りTEUをシェア比率で再配分
                future_share = ba["share"] / future_shares_sum
                pred_teu = int(remaining_teu * future_share + 0.5)
                # 直近の週平均実績を下限とする
                n_actual_weeks = len([x for x in base_alloc if x["has_actual"]])
                if n_actual_weeks > 0:
                    avg_actual = actual_sum / n_actual_weeks
                    pred_teu = max(pred_teu, int(avg_actual * 0.85))
            else:
                pred_teu = ba["base_teu"]

            weekly_pred[wk] = {
                "teu": pred_teu,
                "cm1_per_teu": m_pred["cm1_per_teu"],
            }

    # ── 補正②: 週間TEU予測の平滑化 (先週との急激な乖離を緩和) ──
    sorted_wks = sorted(weekly_pred.keys())
    if len(sorted_wks) >= 3:
        teu_vals = [weekly_pred[wk]["teu"] for wk in sorted_wks]
        smoothed = list(teu_vals)
        for j in range(1, len(teu_vals) - 1):
            smoothed[j] = int((teu_vals[j - 1] + teu_vals[j] * 2 + teu_vals[j + 1]) / 4 + 0.5)
        for j, wk in enumerate(sorted_wks):
            # 平滑化後も実績を下回らないようクランプ
            # 将来週は実績×1.1、過去/当週は実績そのものが下限
            actual_val = weekly_actuals.get(wk, 0)
            w_info = next((w for w in weeks_3m if w["week_key"] == wk), None)
            if w_info and actual_val > 0:
                ws = w_info.get("week_start")
                ws_date = date.fromisoformat(ws) if ws else None
                if ws_date and ws_date > ref:
                    actual_floor = int(actual_val * 1.1 + 0.5)
                else:
                    actual_floor = actual_val
            else:
                actual_floor = actual_val
            weekly_pred[wk]["teu"] = max(smoothed[j], actual_floor)

    # ── 週間 CM1/TEU を線形補間 (月をまたいで滑らかに変化) ──
    # 全週をソートして、月の中央位置に月間CM1値を配置 → 週ごとに補間
    sorted_weeks = sorted(weekly_pred.keys())
    if len(sorted_weeks) >= 2:
        # 各週の月間CM1値と月内位置を取得
        week_cm1_anchors = []  # [(week_idx, cm1_value)]
        for ym_str, cm1_val in _cm1_by_month_idx.items():
            # この月の週のindexを取得
            month_week_idxs = [
                i for i, wk in enumerate(sorted_weeks)
                if any(w["ym"] == ym_str and w["week_key"] == wk
                       for w in weeks_3m)
            ]
            if month_week_idxs:
                mid_idx = month_week_idxs[len(month_week_idxs) // 2]
                week_cm1_anchors.append((mid_idx, cm1_val))

        if len(week_cm1_anchors) >= 2:
            week_cm1_anchors.sort(key=lambda x: x[0])
            for i, wk in enumerate(sorted_weeks):
                # アンカー間で線形補間
                if i <= week_cm1_anchors[0][0]:
                    cm1_v = week_cm1_anchors[0][1]
                elif i >= week_cm1_anchors[-1][0]:
                    cm1_v = week_cm1_anchors[-1][1]
                else:
                    # 前後のアンカーを見つけて補間
                    for j in range(len(week_cm1_anchors) - 1):
                        a_idx, a_val = week_cm1_anchors[j]
                        b_idx, b_val = week_cm1_anchors[j + 1]
                        if a_idx <= i <= b_idx:
                            t = (i - a_idx) / (b_idx - a_idx) if b_idx != a_idx else 0
                            cm1_v = a_val + t * (b_val - a_val)
                            break
                    else:
                        cm1_v = weekly_pred[wk]["cm1_per_teu"]
                weekly_pred[wk]["cm1_per_teu"] = round(cm1_v)

    return {"monthly": monthly_pred, "weekly": weekly_pred}


@app.route("/api/predict")
@require_login
def api_predict():
    """4年分BQデータの季節性分析に基づく将来予測値を返す"""
    today = date.today()
    area = request.args.get("area", "ALL")
    meeting_week = request.args.get("meeting_week", "")
    try:
        result = _build_prediction(area, today, meeting_week)
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── テンプレートデータ API ──────────────────────────────────
# 荷主履歴クエリ (②New Customer / ③Regain 判定用: 過去18か月の荷主別TEU)
SHIPPER_HISTORY_QUERY = """
WITH raw01 AS (
  SELECT Booking_No_, ETD, Booking_Shipper,
    CASE
      WHEN POL IN ("YOK","TYO","CHB","SMZ","NGO","THS","YKK","OSA","UKB") AND CTR = "KR" THEN "JPC_KR"
      WHEN POL NOT IN ("YOK","TYO","CHB","SMZ","NGO","THS","YKK","OSA","UKB") AND CTR = "KR" THEN "JPN_KR"
      ELSE CTR
    END AS AREA,
    SUM(TEU) AS TEU
  FROM `updated_tables.score_dailybooking`
  WHERE ETD >= DATE_SUB(CURRENT_DATE(), INTERVAL 18 MONTH)
  GROUP BY 1,2,3,4
),
raw02 AS (
  SELECT DISTINCT BKG_No, BL_No
  FROM `updated_tables.lifting_detail`
  WHERE FORMAT_DATE('%Y-%m', DATE(Year, Month, 1)) >= FORMAT_DATE('%Y-%m', DATE_SUB(CURRENT_DATE(), INTERVAL 18 MONTH))
),
raw03 AS (
  SELECT DISTINCT FORMAT_DATE('%Y-%m', DATE(Year, MON, 1)) AS YearMonth,
    B_L_No_ AS BL_No, TEU, CM1
  FROM `updated_tables.lpa_revised_monthly`
  WHERE NOT (O_F IS NULL AND CM1 IS NULL)
)
SELECT FORMAT_DATE('%Y-%m', raw01.ETD) AS YearMonth,
  raw01.AREA, raw01.Booking_Shipper,
  SUM(COALESCE(raw03.TEU, raw01.TEU)) AS TEU
FROM raw01
LEFT JOIN raw02 ON raw01.Booking_No_ = raw02.BKG_No
LEFT JOIN raw03 ON raw02.BL_No = raw03.BL_No
GROUP BY 1,2,3
ORDER BY YearMonth
"""

_shipper_hist_df: pd.DataFrame | None = None
_shipper_hist_loaded_at: float = 0
_shipper_hist_lock = threading.Lock()
_SHIPPER_HIST_CACHE_SEC = 3600  # 1時間

def _fetch_shipper_history() -> pd.DataFrame:
    """荷主履歴データをBQから取得 (キャッシュ付き)"""
    global _shipper_hist_df, _shipper_hist_loaded_at
    import time
    now = time.time()
    with _shipper_hist_lock:
        if _shipper_hist_df is not None and (now - _shipper_hist_loaded_at) < _SHIPPER_HIST_CACHE_SEC:
            return _shipper_hist_df
    df = bq_client.query(SHIPPER_HISTORY_QUERY).to_dataframe()
    df["TEU"] = pd.to_numeric(df["TEU"], errors="coerce").fillna(0)
    print(f"Shipper History BQ: {len(df):,} rows")
    with _shipper_hist_lock:
        _shipper_hist_df = df
        _shipper_hist_loaded_at = now
    return df


def _filter_area(df: pd.DataFrame, area: str) -> pd.DataFrame:
    """共通エリアフィルタ (PODベースのサブエリア対応)"""
    if area in SUB_AREA_DEFS:
        parent, pods, include = SUB_AREA_DEFS[area]
        sub = df[df["AREA"] == parent].copy()
        if "POD" in sub.columns:
            if include:
                return sub[sub["POD"].isin(pods)]
            else:
                return sub[~sub["POD"].isin(pods)]
        return sub  # POD列がない場合は親エリア全体
    if area == "KR":
        return df[df["AREA"].isin(["JPC_KR", "JPN_KR"])].copy()
    elif area != "ALL":
        return df[df["AREA"] == area].copy()
    return df.copy()

def _filter_area_no_pod(df: pd.DataFrame, area: str) -> pd.DataFrame:
    """予測用フィルタ (POD列があればサブエリアも対応)"""
    if area in SUB_AREA_DEFS and "POD" in df.columns:
        parent, pod_list, include = SUB_AREA_DEFS[area]
        parent_df = df[df["AREA"] == parent]
        if include:
            return parent_df[parent_df["POD"].isin(pod_list)].copy()
        else:
            return parent_df[~parent_df["POD"].isin(pod_list)].copy()
    parent = _sub_area_parent(area)
    if parent and "POD" not in df.columns:
        area = parent
    if area == "KR":
        return df[df["AREA"].isin(["JPC_KR", "JPN_KR"])].copy()
    elif area != "ALL":
        return df[df["AREA"] == area].copy()
    return df.copy()


def build_template_data(area: str, df: pd.DataFrame, today: date,
                        meeting_week: str = "") -> dict:
    """プレビュー用テンプレートデータを全種類まとめて算出"""
    ref = _ref_date_from_meeting_week(meeting_week, today)
    df_area = _filter_area(df, area)

    if df_area.empty:
        return {}

    ref_ym = f"{ref.year}-{ref.month:02d}"

    # ── 共通: 月別荷主TEU集計 ──────────────────────────
    # 過去5か月 + 当月 + 来月
    months_7 = _cal_months(ref.year, ref.month, 5, 1)
    yms_7 = [f"{y}-{m:02d}" for y, m in months_7]
    df_range = df_area[df_area["bq_ym"].isin(yms_7)]

    # 荷主 × 月別 TEU/CM1
    shipper_monthly = df_range.groupby(["Booking_Shipper", "bq_ym"]).agg(
        TEU=("TEU", "sum"), CM1=("CM1", "sum")
    ).reset_index()

    result = {}

    # ──────────────────────────────────────────────────
    # ① 増減荷主 TOP3 (当月/翌月 × 増加/減少 = 4ブロック)
    # ──────────────────────────────────────────────────
    try:
        import calendar

        # 過去3か月 (当月除く) の月平均 TEU
        months_avg3 = _cal_months(ref.year, ref.month, 3, 0)[:-1]
        yms_avg3 = [f"{y}-{m:02d}" for y, m in months_avg3]
        curr_ym = ref_ym
        next_m = _cal_months(ref.year, ref.month, 0, 1)[-1]
        next_ym = f"{next_m[0]}-{next_m[1]:02d}"

        df_avg3 = shipper_monthly[shipper_monthly["bq_ym"].isin(yms_avg3)]
        # Shipper 単位の 3M平均 TEU（常に3で割る）
        avg3 = (df_avg3.groupby("Booking_Shipper")["TEU"].sum() / 3).rename("avg3_teu")

        # Shipper × POL × DLY 組み合わせの 3M平均 TEU / CM1
        df_detail_avg3 = df_area[df_area["bq_ym"].isin(yms_avg3)]
        avg3_combo = (df_detail_avg3.groupby(["Booking_Shipper", "POL", "DLY"])["TEU"].sum() / 3).rename("avg3_teu")
        avg3_combo_cm1 = (df_detail_avg3.groupby(["Booking_Shipper", "POL", "DLY"])["CM1"].sum() / 3).rename("avg3_cm1")
        avg3_combo_teu_wcm1 = (df_detail_avg3.groupby(["Booking_Shipper", "POL", "DLY"])["TEU_with_cm1"].sum() / 3).rename("avg3_teu_wcm1")

        def _ym_label(ym_str):
            y, m = ym_str.split("-")
            return f"{y} {calendar.month_abbr[int(m)]}"

        avg3_labels = [_ym_label(ym) for ym in yms_avg3]
        avg3_range = f"{avg3_labels[0]} - {avg3_labels[-1]}" if avg3_labels else ""

        # ── (A) Shipper 単位 Top3 ──
        def _top_pol_dly(shipper_name, target_ym, direction="increase"):
            """荷主ごとに、3M平均vs対象月で最も増加/減少したPOL-DLYコンボを返す
            direction: "increase" → 最大増加航路, "decrease" → 最大減少航路"""
            try:
                sub_r = df_area[(df_area["Booking_Shipper"] == shipper_name) & (df_area["bq_ym"] == target_ym)]
                sub_a = df_detail_avg3[df_detail_avg3["Booking_Shipper"] == shipper_name]
                recent_combo = sub_r.groupby(["POL", "DLY"])["TEU"].sum()
                avg3_combo_s = sub_a.groupby(["POL", "DLY"])["TEU"].sum() / 3
                all_idx = recent_combo.index.union(avg3_combo_s.index)
                if len(all_idx) == 0:
                    return None
                diff_s = recent_combo.reindex(all_idx, fill_value=0) - avg3_combo_s.reindex(all_idx, fill_value=0)
                if direction == "increase":
                    filtered = diff_s[diff_s > 0]
                    if filtered.empty:
                        best_idx = diff_s.abs().idxmax()  # フォールバック
                    else:
                        best_idx = filtered.idxmax()
                else:
                    filtered = diff_s[diff_s < 0]
                    if filtered.empty:
                        best_idx = diff_s.abs().idxmax()
                    else:
                        best_idx = filtered.idxmin()
                pol, dly = best_idx
                return {"pol": pol, "dly": dly,
                        "avg3_teu": round(float(avg3_combo_s.get(best_idx, 0))),
                        "recent_teu": round(float(recent_combo.get(best_idx, 0))),
                        "diff": round(float(diff_s[best_idx]))}
            except Exception:
                return None

        def _build_month_block(target_ym):
            """Shipper 単位で 3M平均 vs 対象月 の増減 Top3"""
            df_target = shipper_monthly[shipper_monthly["bq_ym"] == target_ym]
            target_teu = df_target.groupby("Booking_Shipper")["TEU"].sum().rename("recent_teu")
            merged = pd.DataFrame({"avg3_teu": avg3, "recent_teu": target_teu}).fillna(0)
            merged["diff"] = merged["recent_teu"] - merged["avg3_teu"]

            top_inc = merged.nlargest(3, "diff")
            top_inc = top_inc[top_inc["diff"] > 0]
            top_dec = merged.nsmallest(3, "diff")
            top_dec = top_dec[top_dec["diff"] < 0]

            def _rows(top_df, direction="increase"):
                return [{"shipper": s,
                         "avg3_teu": round(float(r["avg3_teu"])),
                         "recent_teu": round(float(r["recent_teu"])),
                         "diff": round(float(r["diff"])),
                         "top_combo": _top_pol_dly(s, target_ym, direction)}
                        for s, r in top_df.iterrows()]

            meta = {"base_months": yms_avg3, "target_month": target_ym,
                    "recent_label": _ym_label(target_ym), "avg3_range": avg3_range}
            return {"items": _rows(top_inc, "increase"), **meta}, {"items": _rows(top_dec, "decrease"), **meta}

        # ── (B) Shipper × POL × DLY 組み合わせ Top3 ──
        def _build_combo_block(target_ym):
            """Shipper×POL×DLY 組み合わせで 3M平均 vs 対象月 の増減 Top3"""
            df_target = df_area[df_area["bq_ym"] == target_ym]
            target_combo = df_target.groupby(["Booking_Shipper", "POL", "DLY"])["TEU"].sum().rename("recent_teu")
            target_combo_cm1 = df_target.groupby(["Booking_Shipper", "POL", "DLY"])["CM1"].sum().rename("recent_cm1")
            target_combo_teu_wcm1 = df_target.groupby(["Booking_Shipper", "POL", "DLY"])["TEU_with_cm1"].sum().rename("recent_teu_wcm1")
            merged = pd.DataFrame({
                "avg3_teu": avg3_combo,
                "recent_teu": target_combo,
                "avg3_cm1": avg3_combo_cm1,
                "avg3_teu_wcm1": avg3_combo_teu_wcm1,
                "recent_cm1": target_combo_cm1,
                "recent_teu_wcm1": target_combo_teu_wcm1,
            }).fillna(0)
            merged["diff"] = merged["recent_teu"] - merged["avg3_teu"]
            # CM1/TEU: CM1÷TEU_with_cm1 (運賃紐づきTEUのみ分母)
            merged["avg3_cm1t"] = (merged["avg3_cm1"] / merged["avg3_teu_wcm1"].replace(0, float("nan"))).round(0).fillna(0)
            merged["recent_cm1t"] = (merged["recent_cm1"] / merged["recent_teu_wcm1"].replace(0, float("nan"))).round(0).fillna(0)
            merged["cm1t_diff"] = merged["recent_cm1t"] - merged["avg3_cm1t"]

            top_inc = merged.nlargest(5, "diff")
            top_inc = top_inc[top_inc["diff"] > 0]
            top_dec = merged.nsmallest(5, "diff")
            top_dec = top_dec[top_dec["diff"] < 0]

            def _rows(top_df):
                rows = []
                for idx, row in top_df.iterrows():
                    shipper, pol, dly = idx
                    r = {"shipper": shipper, "pol": pol, "dly": dly,
                         "avg3_teu": round(float(row["avg3_teu"])),
                         "recent_teu": round(float(row["recent_teu"])),
                         "diff": round(float(row["diff"])),
                         "avg3_cm1t": round(float(row["avg3_cm1t"])),
                         "recent_cm1t": round(float(row["recent_cm1t"])),
                         "cm1t_diff": round(float(row["cm1t_diff"]))}
                    rows.append(r)
                return rows

            meta = {"base_months": yms_avg3, "target_month": target_ym,
                    "recent_label": _ym_label(target_ym), "avg3_range": avg3_range}
            return {"items": _rows(top_inc), **meta}, {"items": _rows(top_dec), **meta}

        # Shipper 単位
        inc_curr, dec_curr = _build_month_block(curr_ym)
        inc_next, dec_next = _build_month_block(next_ym)
        result["shipper_increase_curr"] = inc_curr
        result["shipper_increase_next"] = inc_next
        result["shipper_decrease_curr"] = dec_curr
        result["shipper_decrease_next"] = dec_next

        # Shipper × POL × DLY 組み合わせ
        cinc_curr, cdec_curr = _build_combo_block(curr_ym)
        cinc_next, cdec_next = _build_combo_block(next_ym)
        result["combo_increase_curr"] = cinc_curr
        result["combo_increase_next"] = cinc_next
        result["combo_decrease_curr"] = cdec_curr
        result["combo_decrease_next"] = cdec_next

    except Exception as e:
        print(f"Template shipper_change error: {e}")
        result["shipper_increase_curr"] = {"items": []}
        result["shipper_increase_next"] = {"items": []}
        result["shipper_decrease_curr"] = {"items": []}
        result["shipper_decrease_next"] = {"items": []}
        result["combo_increase_curr"] = {"items": []}
        result["combo_increase_next"] = {"items": []}
        result["combo_decrease_curr"] = {"items": []}
        result["combo_decrease_next"] = {"items": []}

    # ──────────────────────────────────────────────────
    # ② New Customer / ③ Regain Customer
    # ──────────────────────────────────────────────────
    try:
        hist_df = _fetch_shipper_history()
        hist_area = _filter_area(hist_df, area)

        # 月リスト作成
        months_12_ago = _cal_months(ref.year, ref.month, 12, 0)[:-1]  # 当月除く過去12か月
        months_6_ago = _cal_months(ref.year, ref.month, 6, 0)[:-1]   # 当月除く過去6か月
        months_7to12 = [m for m in months_12_ago if m not in months_6_ago]
        yms_12 = [f"{y}-{m:02d}" for y, m in months_12_ago]
        yms_6 = [f"{y}-{m:02d}" for y, m in months_6_ago]
        yms_7to12 = [f"{y}-{m:02d}" for y, m in months_7to12]
        yms_recent = [curr_ym, next_ym] if next_ym in hist_area["YearMonth"].values else [curr_ym]
        # BASE_QUERY の当月・来月荷主
        recent_shippers = df_area[df_area["bq_ym"].isin([curr_ym, next_ym])]
        recent_shipper_teu = recent_shippers.groupby("Booking_Shipper")["TEU"].sum()
        active_now = set(recent_shipper_teu[recent_shipper_teu > 0].index)

        # 過去12か月の荷主TEU (履歴クエリから)
        hist_12 = hist_area[hist_area["YearMonth"].isin(yms_12)]
        shippers_12 = set(hist_12.groupby("Booking_Shipper")["TEU"].sum()
                         .pipe(lambda s: s[s > 0]).index)

        # 過去6か月の荷主TEU
        hist_6 = hist_area[hist_area["YearMonth"].isin(yms_6)]
        shippers_6 = set(hist_6.groupby("Booking_Shipper")["TEU"].sum()
                        .pipe(lambda s: s[s > 0]).index)

        # 7-12か月前の荷主TEU
        hist_7to12 = hist_area[hist_area["YearMonth"].isin(yms_7to12)]
        shippers_7to12 = set(hist_7to12.groupby("Booking_Shipper")["TEU"].sum()
                            .pipe(lambda s: s[s > 0]).index)

        # 荷主ごとのトップ航路を求めるヘルパー
        def _top_route_for_shipper(shipper_name):
            """当月＋来月の荷量が最大のPOL-DLYコンボを返す"""
            try:
                sub = df_area[(df_area["Booking_Shipper"] == shipper_name) & (df_area["bq_ym"].isin([curr_ym, next_ym]))]
                if sub.empty:
                    return None
                combo = sub.groupby(["POL", "DLY"])["TEU"].sum()
                if combo.empty:
                    return None
                best = combo.idxmax()
                return {"pol": best[0], "dly": best[1], "teu": round(float(combo[best]))}
            except Exception:
                return None

        # ② New Customer: 過去12か月TEU=0 かつ 今活動中
        new_customers_set = active_now - shippers_12
        new_cust_list = []
        for s in new_customers_set:
            teu = float(recent_shipper_teu.get(s, 0))
            new_cust_list.append({"shipper": s, "teu": round(teu), "top_route": _top_route_for_shipper(s)})
        new_cust_list.sort(key=lambda x: -x["teu"])
        overflow_new = len(new_cust_list) - 3 if len(new_cust_list) > 3 else 0
        overflow_new_teu = sum(x["teu"] for x in new_cust_list[3:]) if overflow_new > 0 else 0

        result["new_customer"] = {
            "customers": new_cust_list[:3],
            "overflow_count": overflow_new,
            "overflow_teu": round(overflow_new_teu),
            "total_count": len(new_cust_list),
        }

        # ③ Regain: 7-12か月前に活動, 過去6か月は0, 今活動中
        regain_set = (active_now & shippers_7to12) - shippers_6
        regain_list = []
        for s in regain_set:
            teu = float(recent_shipper_teu.get(s, 0))
            regain_list.append({"shipper": s, "teu": round(teu), "top_route": _top_route_for_shipper(s)})
        regain_list.sort(key=lambda x: -x["teu"])
        overflow_regain = len(regain_list) - 3 if len(regain_list) > 3 else 0
        overflow_regain_teu = sum(x["teu"] for x in regain_list[3:]) if overflow_regain > 0 else 0

        result["regain_customer"] = {
            "customers": regain_list[:3],
            "overflow_count": overflow_regain,
            "overflow_teu": round(overflow_regain_teu),
            "total_count": len(regain_list),
        }
    except Exception as e:
        print(f"Template new/regain error: {e}")
        import traceback; traceback.print_exc()
        result["new_customer"] = {"customers": [], "overflow_count": 0, "overflow_teu": 0, "total_count": 0}
        result["regain_customer"] = {"customers": [], "overflow_count": 0, "overflow_teu": 0, "total_count": 0}

    # ──────────────────────────────────────────────────
    # ④ POL数 (Monthly)
    # ──────────────────────────────────────────────────
    try:
        months_6 = _cal_months(ref.year, ref.month, 5, 0)
        yms_6_list = [f"{y}-{m:02d}" for y, m in months_6]
        pol_data = []
        for ym in yms_6_list:
            sub = df_area[df_area["bq_ym"] == ym]
            pol_count = int(sub["POL"].nunique())
            pol_data.append({"ym": ym, "pol_count": pol_count})
        result["pol_count"] = pol_data
    except Exception as e:
        print(f"Template pol_count error: {e}")
        result["pol_count"] = []

    # ──────────────────────────────────────────────────
    # ⑤ Booking件数 (Monthly & Weekly)
    # ──────────────────────────────────────────────────
    MAIN_POLS = {"TYO", "YOK", "NGO", "OSA", "UKB"}
    try:
        # Monthly (Main / Local 内訳)
        months_6 = _cal_months(ref.year, ref.month, 5, 0)
        yms_6_list = [f"{y}-{m:02d}" for y, m in months_6]
        bkg_monthly = []
        for ym in yms_6_list:
            sub = df_area[df_area["bq_ym"] == ym]
            sub_main = sub[sub["POL"].isin(MAIN_POLS)]
            sub_local = sub[~sub["POL"].isin(MAIN_POLS)]
            bkg_monthly.append({
                "ym": ym,
                "count": int(sub["Booking_No_"].nunique()),
                "main_count": int(sub_main["Booking_No_"].nunique()),
                "local_count": int(sub_local["Booking_No_"].nunique()),
            })

        # Weekly (3か月分)
        weeks_3m = get_3month_weeks(ref)
        bkg_weekly = []
        for w in weeks_3m:
            sub = df_area[df_area["week_key"] == w["week_key"]]
            bkg_count = int(sub["Booking_No_"].nunique())
            bkg_weekly.append({
                "week_key": w["week_key"],
                "week_label": f"W{w['week']}",
                "ym": w["ym"],
                "count": bkg_count,
            })
        result["booking_count"] = {"monthly": bkg_monthly, "weekly": bkg_weekly}
    except Exception as e:
        print(f"Template booking_count error: {e}")
        result["booking_count"] = {"monthly": [], "weekly": []}

    # ──────────────────────────────────────────────────
    # ⑥ 営業マン寄与度
    # ──────────────────────────────────────────────────
    try:
        months_3 = _cal_months(ref.year, ref.month, 2, 0)
        sales_data = {}
        for y, m in months_3:
            ym = f"{y}-{m:02d}"
            sub = df_area[df_area["bq_ym"] == ym]
            total_teu = float(sub["TEU"].sum())
            if total_teu <= 0:
                sales_data[ym] = {"items": [], "total_teu": 0}
                continue
            by_sales = sub.groupby("POL_Sales")["TEU"].sum().sort_values(ascending=False)
            items = []
            for sales, teu in by_sales.items():
                items.append({
                    "sales": sales,
                    "teu": round(float(teu)),
                    "pct": round(float(teu) / total_teu * 100, 1),
                })
            sales_data[ym] = {"items": items, "total_teu": round(total_teu)}
        result["sales_contribution"] = sales_data
    except Exception as e:
        print(f"Template sales_contribution error: {e}")
        result["sales_contribution"] = {}

    # ──────────────────────────────────────────────────
    # 古紙荷主リスト
    # 会議≤14日: 過去3か月TEU + 3M平均 + 当月実績 + Gap
    # 会議≥15日: 過去3か月TEU + 3M平均 + 翌月実績 + Gap
    # ──────────────────────────────────────────────────
    KOSHI_SHIPPER_CODES = {
        "SEGJ04", "NIPJ01", "LNAJ01", "SCJJ14", "SAMP04",
        "KPPJ01", "KYUJ11", "TCIJ00", "NTRJ03",
    }
    try:
        # 会議日の14/15判定
        meeting_tue_day = (ref + timedelta(days=2)).day
        if meeting_tue_day <= 14:
            # 過去3か月 = ref月-3, ref月-2, ref月-1  / 対象月 = ref月(当月)
            koshi_past_months = _cal_months(ref.year, ref.month, 3, 0)[:-1]  # 過去3か月
            target_y, target_m = ref.year, ref.month  # 当月
        else:
            # 過去3か月 = ref月-2, ref月-1, ref月  / 対象月 = ref月+1(翌月)
            koshi_past_months = _cal_months(ref.year, ref.month, 2, 0)       # 過去3か月(当月含む)
            tgt = _cal_months(ref.year, ref.month, 0, 1)
            target_y, target_m = tgt[-1]  # 翌月

        koshi_past_yms = [f"{y}-{m:02d}" for y, m in koshi_past_months]
        target_ym = f"{target_y}-{target_m:02d}"
        all_koshi_yms = koshi_past_yms + [target_ym]

        df_koshi = df_area[
            (df_area["BKG_Shipper_code"].isin(KOSHI_SHIPPER_CODES)) &
            (df_area["bq_ym"].isin(all_koshi_yms))
        ]

        # 荷主ごとに集計
        all_shippers = df_koshi.groupby("Booking_Shipper")
        koshi_items = []
        for shipper_name, grp in all_shippers:
            past_months_data = []
            past_total = 0
            for ym in koshi_past_yms:
                sub = grp[grp["bq_ym"] == ym]
                teu = round(float(sub["TEU"].sum()))
                past_months_data.append({"ym": ym, "teu": teu})
                past_total += teu
            avg_3m = round(past_total / 3) if len(koshi_past_yms) == 3 else 0
            # 対象月の実績
            target_sub = grp[grp["bq_ym"] == target_ym]
            target_teu = round(float(target_sub["TEU"].sum()))
            gap = target_teu - avg_3m
            koshi_items.append({
                "shipper": shipper_name,
                "past_months": past_months_data,
                "avg_3m": avg_3m,
                "target_teu": target_teu,
                "gap": gap,
            })

        # Gap降順でソート
        koshi_items.sort(key=lambda x: x["gap"], reverse=True)

        result["koshi_shipper"] = {
            "past_yms": koshi_past_yms,
            "target_ym": target_ym,
            "is_before_15": meeting_tue_day <= 14,
            "items": koshi_items,
        }
    except Exception as e:
        print(f"Template koshi_shipper error: {e}")
        result["koshi_shipper"] = {}

    # ──────────────────────────────────────────────────
    # ⑦ CM1レンジ分析 (Profitability Segmentation)
    # ──────────────────────────────────────────────────
    try:
        # 直近3か月の荷主別CM1/TEU
        months_3_range = _cal_months(ref.year, ref.month, 2, 0)
        yms_3 = [f"{y}-{m:02d}" for y, m in months_3_range]

        cm1_range_data = {}
        for ym in yms_3:
            sub = df_area[df_area["bq_ym"] == ym]
            shipper_agg = sub.groupby("Booking_Shipper").agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum")
            ).reset_index()
            shipper_agg = shipper_agg[shipper_agg["TEU"] > 0]
            shipper_agg["CM1_per_TEU"] = shipper_agg["CM1"] / shipper_agg["TEU_with_cm1"].replace(0, float("nan"))

            if shipper_agg.empty:
                cm1_range_data[ym] = {"high": {}, "mid": {}, "low": {}}
                continue

            # 四分位でHigh/Mid/Low分類
            q75 = float(shipper_agg["CM1_per_TEU"].quantile(0.75))
            q25 = float(shipper_agg["CM1_per_TEU"].quantile(0.25))

            high = shipper_agg[shipper_agg["CM1_per_TEU"] >= q75]
            mid = shipper_agg[(shipper_agg["CM1_per_TEU"] >= q25) & (shipper_agg["CM1_per_TEU"] < q75)]
            low = shipper_agg[shipper_agg["CM1_per_TEU"] < q25]

            total_teu = float(shipper_agg["TEU"].sum())

            def _seg_info(seg_df):
                teu = float(seg_df["TEU"].sum())
                cm1 = float(seg_df["CM1"].sum())
                teu_wcm1 = float(seg_df["TEU_with_cm1"].sum())
                return {
                    "teu": round(teu),
                    "pct": round(teu / total_teu * 100, 1) if total_teu > 0 else 0,
                    "shipper_count": len(seg_df),
                    "cm1_per_teu": round(cm1 / teu_wcm1) if teu_wcm1 > 0 else 0,
                }

            cm1_range_data[ym] = {
                "high": _seg_info(high),
                "mid": _seg_info(mid),
                "low": _seg_info(low),
                "q75": round(q75),
                "q25": round(q25),
            }
        result["cm1_range"] = cm1_range_data
    except Exception as e:
        print(f"Template cm1_range error: {e}")
        result["cm1_range"] = {}

    # ──────────────────────────────────────────────────
    # ⑧ Trade Lane ヒートマップ
    # ──────────────────────────────────────────────────
    try:
        months_6 = _cal_months(ref.year, ref.month, 5, 0)
        yms_6_list = [f"{y}-{m:02d}" for y, m in months_6]

        # ALL → CTR列(国コード2文字), エリア別 → DLY列(港コード3文字)
        group_col = "CTR" if area == "ALL" else "DLY"

        df_6m = df_area[df_area["bq_ym"].isin(yms_6_list)]

        # グループ × 月別集計
        lane_agg = df_6m.groupby([group_col, "bq_ym"]).agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum"),
            TEU_with_cm1=("TEU_with_cm1", "sum")
        ).reset_index()

        # 上位10仕向地/国を選定
        top_lanes = (df_6m.groupby(group_col)["TEU"].sum()
                    .sort_values(ascending=False).head(10).index.tolist())

        heatmap = []
        for lane in top_lanes:
            row = {"lane": lane, "months": {}}
            for ym in yms_6_list:
                sub = lane_agg[(lane_agg[group_col] == lane) & (lane_agg["bq_ym"] == ym)]
                if not sub.empty:
                    teu = float(sub["TEU"].iloc[0])
                    cm1 = float(sub["CM1"].iloc[0])
                    teu_wcm1 = float(sub["TEU_with_cm1"].iloc[0])
                    row["months"][ym] = {
                        "teu": round(teu),
                        "cm1_per_teu": round(cm1 / teu_wcm1) if teu_wcm1 > 0 else 0,
                    }
                else:
                    row["months"][ym] = {"teu": 0, "cm1_per_teu": 0}
            heatmap.append(row)

        result["trade_lane"] = {
            "group_by": "CTR" if area == "ALL" else "DLY",
            "months": yms_6_list,
            "data": heatmap,
        }
    except Exception as e:
        print(f"Template trade_lane error: {e}")
        result["trade_lane"] = {"group_by": "", "months": [], "data": []}

    # ──────────────────────────────────────────────────
    # ⑪ CM1/TEU ウォーターフォール (前月比要因分解)
    # ──────────────────────────────────────────────────
    try:
        prev_m = _cal_months(ref.year, ref.month, 1, 0)[0]
        prev_ym = f"{prev_m[0]}-{prev_m[1]:02d}"

        df_prev = df_area[df_area["bq_ym"] == prev_ym]
        df_curr = df_area[df_area["bq_ym"] == ref_ym]

        # 荷主別集計
        prev_agg = df_prev.groupby("Booking_Shipper").agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum")).reset_index()
        curr_agg = df_curr.groupby("Booking_Shipper").agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum")).reset_index()

        prev_total_teu = float(prev_agg["TEU"].sum())
        prev_total_cm1 = float(prev_agg["CM1"].sum())
        curr_total_teu = float(curr_agg["TEU"].sum())
        curr_total_cm1 = float(curr_agg["CM1"].sum())

        prev_cm1t = round(prev_total_cm1 / prev_total_teu) if prev_total_teu > 0 else 0
        curr_cm1t = round(curr_total_cm1 / curr_total_teu) if curr_total_teu > 0 else 0

        # Mix効果: 前月の単価 × (当月構成比 - 前月構成比) の荷主合計
        # Rate効果: 当月構成比 × (当月単価 - 前月単価) の荷主合計
        all_shippers = set(prev_agg["Booking_Shipper"].tolist() + curr_agg["Booking_Shipper"].tolist())

        mix_effect = 0
        rate_effect = 0
        for s in all_shippers:
            p = prev_agg[prev_agg["Booking_Shipper"] == s]
            c = curr_agg[curr_agg["Booking_Shipper"] == s]

            p_teu = float(p["TEU"].iloc[0]) if len(p) > 0 else 0
            p_cm1 = float(p["CM1"].iloc[0]) if len(p) > 0 else 0
            c_teu = float(c["TEU"].iloc[0]) if len(c) > 0 else 0
            c_cm1 = float(c["CM1"].iloc[0]) if len(c) > 0 else 0

            p_share = p_teu / prev_total_teu if prev_total_teu > 0 else 0
            c_share = c_teu / curr_total_teu if curr_total_teu > 0 else 0
            p_rate = p_cm1 / p_teu if p_teu > 0 else 0
            c_rate = c_cm1 / c_teu if c_teu > 0 else 0

            # Mix = 前月単価 × (当月構成比 - 前月構成比)
            mix_effect += p_rate * (c_share - p_share)
            # Rate = 当月構成比 × (当月単価 - 前月単価)
            rate_effect += c_share * (c_rate - p_rate)

        # Volume効果 (残差)
        total_change = curr_cm1t - prev_cm1t
        volume_effect = total_change - mix_effect - rate_effect

        result["cm1_waterfall"] = {
            "prev_ym": prev_ym,
            "curr_ym": ref_ym,
            "prev_cm1t": prev_cm1t,
            "curr_cm1t": curr_cm1t,
            "mix_effect": round(mix_effect),
            "rate_effect": round(rate_effect),
            "volume_effect": round(volume_effect),
            "total_change": curr_cm1t - prev_cm1t,
        }
    except Exception as e:
        print(f"Template cm1_waterfall error: {e}")
        result["cm1_waterfall"] = {}

    return result


@app.route("/api/template-data")
@require_login
def api_template_data():
    """プレビュー用テンプレートデータを返す ?area=KR&meeting_week=2026-W14"""
    today = date.today()
    df = get_bq_df()
    area = request.args.get("area", "ALL")
    meeting_week = request.args.get("meeting_week", "")
    try:
        result = build_template_data(area, df, today, meeting_week)
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── テンプレート設定 API ─────────────────────────────────
@app.route("/api/template-definitions")
@require_login
def api_template_definitions():
    """利用可能なテンプレート定義一覧を返す"""
    return jsonify(TEMPLATE_DEFINITIONS)



def _get_default_template_config() -> list:
    """デフォルトのテンプレート設定を生成"""
    return [
        {"id": t["id"], "enabled": t["default_on"],
         "col_span": t["col_span"], "height": 0, "order": i}
        for i, t in enumerate(TEMPLATE_DEFINITIONS)
    ]


@app.route("/api/template-config")
@require_login
def api_get_template_config():
    """テンプレート設定を取得 ?week_key=...&area=..."""
    week_key = request.args.get("week_key", "")
    area = request.args.get("area", "ALL")
    doc_id = f"{week_key}_{area}" if week_key else f"default_{area}"

    doc = db.collection(FS_TEMPLATE_CONFIG).document(doc_id).get()
    if doc.exists:
        d = doc.to_dict()
        return jsonify({
            "exists": True,
            "blocks": d.get("blocks", []),
            "updated": d.get("updated"),
            "updated_by": d.get("updated_by"),
        })

    # Area固定スタイルがあればそれを返す
    fixed_doc = db.collection(FS_AREA_FIXED_STYLE).document(area).get()
    if fixed_doc.exists:
        fd = fixed_doc.to_dict()
        return jsonify({
            "exists": False,
            "from_fixed": True,
            "blocks": fd.get("blocks", []),
        })

    # デフォルトテンプレート (★マーク付き) があればそれを返す
    default_blocks = _get_default_style_template_blocks()
    if default_blocks is not None:
        return jsonify({
            "exists": False,
            "from_default_tpl": True,
            "blocks": default_blocks,
        })

    # コード上のデフォルト設定
    return jsonify({
        "exists": False,
        "blocks": _get_default_template_config(),
    })


@app.route("/api/template-config", methods=["POST"])
@require_editor
def api_save_template_config():
    """テンプレート設定を保存"""
    data = request.json or {}
    week_key = data.get("week_key", "")
    area = data.get("area", "ALL")
    blocks = data.get("blocks", [])
    user = get_current_user()

    doc_id = f"{week_key}_{area}" if week_key else f"default_{area}"
    db.collection(FS_TEMPLATE_CONFIG).document(doc_id).set({
        "week_key": week_key,
        "area": area,
        "blocks": blocks,
        "updated": datetime.now(tz=JST).isoformat(),
        "updated_by": user["email"] if user else "unknown",
    })
    return jsonify({"ok": True})


@app.route("/api/template-config/propagate", methods=["POST"])
@require_editor
def api_propagate_template_config():
    """テンプレート設定を他のArea/週に適用
    mode: "area_one" | "area_all" | "week_same_area" | "week_all"
    """
    data = request.json or {}
    source_week = data.get("week_key", "")
    source_area = data.get("area", "ALL")
    blocks = data.get("blocks", [])
    mode = data.get("mode", "")
    target_area = data.get("target_area", "")
    user = get_current_user()

    if not blocks or not mode:
        return jsonify({"error": "blocks and mode required"}), 400

    now_jst = datetime.now(tz=JST).isoformat()

    # 既存の設定がある週/エリアは上書きしない
    existing_docs = set()
    for doc in db.collection(FS_TEMPLATE_CONFIG).stream():
        existing_docs.add(doc.id)

    # 対象一覧を決定
    targets = []
    df = get_bq_df()
    all_areas = get_all_areas(df)

    if mode == "area_one" and target_area:
        # 同じ週の指定エリアに適用
        doc_id = f"{source_week}_{target_area}"
        if doc_id not in existing_docs:
            targets.append((source_week, target_area))

    elif mode == "area_all":
        # 同じ週の全エリアに適用
        for a in all_areas:
            doc_id = f"{source_week}_{a}"
            if doc_id not in existing_docs and a != source_area:
                targets.append((source_week, a))

    elif mode == "week_same_area":
        # 全週の同じエリアに適用 (既存設定がない週のみ)
        # スナップショット一覧から週を取得
        snap_docs = db.collection(FS_SNAPSHOTS).stream()
        for sd in snap_docs:
            wk = sd.to_dict().get("week_key", sd.id)
            doc_id = f"{wk}_{source_area}"
            if doc_id not in existing_docs and wk != source_week:
                targets.append((wk, source_area))

    elif mode == "week_all":
        # 全週の全エリアに適用 (既存設定がない組み合わせのみ)
        snap_docs = db.collection(FS_SNAPSHOTS).stream()
        week_keys = set()
        for sd in snap_docs:
            week_keys.add(sd.to_dict().get("week_key", sd.id))
        for wk in week_keys:
            for a in all_areas:
                doc_id = f"{wk}_{a}"
                if doc_id not in existing_docs and not (wk == source_week and a == source_area):
                    targets.append((wk, a))

    # バッチ書き込み
    applied = 0
    batch = db.batch()
    for wk, a in targets:
        doc_id = f"{wk}_{a}"
        ref = db.collection(FS_TEMPLATE_CONFIG).document(doc_id)
        batch.set(ref, {
            "week_key": wk,
            "area": a,
            "blocks": blocks,
            "updated": now_jst,
            "updated_by": user["email"] if user else "unknown",
            "propagated_from": f"{source_week}_{source_area}",
        })
        applied += 1
        # Firestore batch limit: 500
        if applied % 450 == 0:
            batch.commit()
            batch = db.batch()
    if applied > 0:
        batch.commit()

    return jsonify({"ok": True, "applied": applied})


# ── スタイルテンプレート (名前付き保存/呼び出し) ───────────
FS_STYLE_TEMPLATES = "meeting_style_templates"

@app.route("/api/style-templates")
@require_login
def api_list_style_templates():
    """保存済みスタイルテンプレート一覧"""
    try:
        docs = db.collection(FS_STYLE_TEMPLATES).stream()
        result = []
        for d in docs:
            data = d.to_dict()
            result.append({
                "id": d.id,
                "name": data.get("name", d.id),
                "created": data.get("created"),
                "created_by": data.get("created_by"),
                "block_count": len(data.get("blocks", [])),
                "is_default": data.get("is_default", False),
            })
        result.sort(key=lambda x: x.get("created") or "", reverse=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/style-templates", methods=["POST"])
@require_editor
def api_save_style_template():
    """現在のスタイルをテンプレートとして保存"""
    data = request.json or {}
    name = data.get("name", "").strip()
    blocks = data.get("blocks", [])
    user = get_current_user()
    if not name:
        return jsonify({"error": "name required"}), 400

    doc_ref = db.collection(FS_STYLE_TEMPLATES).document()
    doc_ref.set({
        "name": name,
        "blocks": blocks,
        "created": datetime.now(tz=JST).isoformat(),
        "created_by": user["email"] if user else "unknown",
    })
    return jsonify({"ok": True, "id": doc_ref.id})

@app.route("/api/style-templates/<tpl_id>")
@require_login
def api_get_style_template(tpl_id):
    """テンプレートの内容を取得"""
    doc = db.collection(FS_STYLE_TEMPLATES).document(tpl_id).get()
    if not doc.exists:
        return jsonify({"error": "not found"}), 404
    data = doc.to_dict()
    return jsonify({"id": doc.id, "name": data.get("name"), "blocks": data.get("blocks", [])})

@app.route("/api/style-templates/<tpl_id>", methods=["DELETE"])
@require_editor
def api_delete_style_template(tpl_id):
    """テンプレートを削除（★DEFAULTは削除不可）"""
    doc = db.collection(FS_STYLE_TEMPLATES).document(tpl_id).get()
    if not doc.exists:
        return jsonify({"error": "not found"}), 404
    data = doc.to_dict()
    if data.get("is_default"):
        return jsonify({"error": "デフォルトテンプレートは削除できません"}), 403
    # 作成者のみ削除可能
    user = get_current_user()
    if data.get("created_by") != (user or {}).get("email"):
        return jsonify({"error": "自分のテンプレートのみ削除できます"}), 403
    db.collection(FS_STYLE_TEMPLATES).document(tpl_id).delete()
    return jsonify({"ok": True})

@app.route("/api/style-templates/<tpl_id>/set-default", methods=["POST"])
@require_editor
def api_set_default_style_template(tpl_id):
    """テンプレートをデフォルトに設定（既存のデフォルトは解除）。作成者のみ"""
    doc = db.collection(FS_STYLE_TEMPLATES).document(tpl_id).get()
    if not doc.exists:
        return jsonify({"error": "not found"}), 404
    user = get_current_user()
    if doc.to_dict().get("created_by") != (user or {}).get("email"):
        return jsonify({"error": "自分のテンプレートのみ設定できます"}), 403
    # 全テンプレートの is_default を解除
    for d in db.collection(FS_STYLE_TEMPLATES).stream():
        if d.to_dict().get("is_default"):
            db.collection(FS_STYLE_TEMPLATES).document(d.id).update({"is_default": False})
    # 指定テンプレートをデフォルトに設定
    db.collection(FS_STYLE_TEMPLATES).document(tpl_id).update({"is_default": True})
    return jsonify({"ok": True})

@app.route("/api/style-templates/<tpl_id>/unset-default", methods=["POST"])
@require_editor
def api_unset_default_style_template(tpl_id):
    """デフォルト設定を解除。作成者のみ"""
    doc = db.collection(FS_STYLE_TEMPLATES).document(tpl_id).get()
    if not doc.exists:
        return jsonify({"error": "not found"}), 404
    user = get_current_user()
    if doc.to_dict().get("created_by") != (user or {}).get("email"):
        return jsonify({"error": "自分のテンプレートのみ変更できます"}), 403
    db.collection(FS_STYLE_TEMPLATES).document(tpl_id).update({"is_default": False})
    return jsonify({"ok": True})

def _get_default_style_template_blocks():
    """is_default=True のスタイルテンプレートの blocks を返す (なければ None)"""
    try:
        for d in db.collection(FS_STYLE_TEMPLATES).stream():
            data = d.to_dict()
            if data.get("is_default"):
                return data.get("blocks", [])
    except Exception:
        pass
    return None


# ── Area固定スタイル ───────────────────────────────────
FS_AREA_FIXED_STYLE = "meeting_area_fixed_style"

@app.route("/api/area-fixed-style")
@require_login
def api_get_area_fixed_style():
    """Areaの固定スタイルを取得 ?area=..."""
    area = request.args.get("area", "")
    if not area:
        return jsonify({"fixed": False})
    doc = db.collection(FS_AREA_FIXED_STYLE).document(area).get()
    if doc.exists:
        data = doc.to_dict()
        return jsonify({"fixed": True, "blocks": data.get("blocks", []), "set_by": data.get("set_by")})
    return jsonify({"fixed": False})

@app.route("/api/area-fixed-style", methods=["POST"])
@require_editor
def api_set_area_fixed_style():
    """Areaの固定スタイルを設定/解除"""
    data = request.json or {}
    area = data.get("area", "")
    blocks = data.get("blocks")  # None = 解除
    user = get_current_user()
    if not area:
        return jsonify({"error": "area required"}), 400

    if blocks is None:
        # 解除
        db.collection(FS_AREA_FIXED_STYLE).document(area).delete()
        return jsonify({"ok": True, "fixed": False})

    db.collection(FS_AREA_FIXED_STYLE).document(area).set({
        "area": area,
        "blocks": blocks,
        "set_by": user["email"] if user else "unknown",
        "updated": datetime.now(tz=JST).isoformat(),
    })
    return jsonify({"ok": True, "fixed": True})


# ── テンプレート設定取得: Area固定スタイルのフォールバック対応 ──
# (既存の api_get_template_config を修正)


# ── スナップショット API ─────────────────────────────────
@app.route("/api/snapshot", methods=["POST"])
@require_editor
def api_save_snapshot():
    """選択週のスナップショットを保存（BQ再取得 + 全エリア分を凍結）"""
    data = request.json or {}
    week_key = data.get("week_key")
    if not week_key:
        return jsonify({"error": "week_key is required"}), 400

    user = get_current_user()

    # BQ データ再取得
    ok = do_refresh_bq(update_firestore=True, refreshed_by=user["email"])
    if not ok:
        return jsonify({"error": "BigQuery更新に失敗しました"}), 500

    df = get_bq_df()
    today = date.today()
    areas = get_all_areas(df)

    # 全エリア分のサマリーを構築（meeting_week で見込みスコープ）
    areas_data = {}
    for a in areas:
        areas_data[a] = build_summary_for_area(a, df, today, meeting_week=week_key)

    now_jst = datetime.now(tz=JST).isoformat()
    snapshot_doc = {
        "week_key": week_key,
        "created_at": now_jst,
        "created_by": user["email"],
        "areas": areas_data,
    }

    # Firestore に保存
    db.collection(FS_SNAPSHOTS).document(week_key).set(snapshot_doc)
    print(f"Snapshot saved: {week_key} by {user['email']} ({len(areas)} areas)")

    info = get_last_refresh()
    return jsonify({
        "ok": True,
        "week_key": week_key,
        "last_refresh": info.get("last_refresh") if info else now_jst,
        "refreshed_by": user["email"],
        "created_at": now_jst,
    })


@app.route("/api/snapshot")
@require_login
def api_get_snapshot():
    """保存済みスナップショットを取得 ?week_key=2026-W14&area=KR"""
    week_key = request.args.get("week_key")
    area = request.args.get("area", "ALL")
    if not week_key:
        return jsonify({"error": "week_key is required"}), 400

    doc = db.collection(FS_SNAPSHOTS).document(week_key).get()
    if not doc.exists:
        return jsonify({"exists": False})

    snap = doc.to_dict()
    area_data = snap.get("areas", {}).get(area, {})
    # サブエリアが古いスナップショットにない場合 → ライブデータで補完
    if not area_data and area in SUB_AREA_DEFS:
        try:
            df = get_bq_df()
            today = date.today()
            area_data = build_summary_for_area(area, df, today, meeting_week=week_key)
        except Exception as e:
            print(f"Sub-area live fallback error: {e}")
            area_data = {}
    return jsonify({
        "exists": True,
        "week_key": week_key,
        "created_at": snap.get("created_at"),
        "created_by": snap.get("created_by"),
        **area_data,
    })


@app.route("/api/snapshot/list")
@require_login
def api_snapshot_list():
    """スナップショットが存在する週一覧を返す"""
    docs = db.collection(FS_SNAPSHOTS).stream()
    snapshots = []
    for doc in docs:
        d = doc.to_dict()
        snapshots.append({
            "week_key": d.get("week_key", doc.id),
            "created_at": d.get("created_at"),
            "created_by": d.get("created_by"),
        })
    snapshots.sort(key=lambda x: x["week_key"])
    return jsonify({"snapshots": snapshots})


def _auto_sum_parent_prospect(parent_area: str, week_key: str, meeting_week: str):
    """子エリアの週間見込みを合算して親エリアに保存"""
    children = AREA_AUTO_SUM.get(parent_area, [])
    total_teu, total_cm1 = 0.0, 0.0
    any_teu, any_cm1 = False, False
    for child in children:
        doc_id = f"{meeting_week}_{week_key}_{child}" if meeting_week else f"{week_key}_{child}"
        doc = db.collection(FS_PROSPECTS).document(doc_id).get()
        if doc.exists:
            d = doc.to_dict()
            if d.get("teu") is not None:
                total_teu += float(d["teu"])
                any_teu = True
            if d.get("cm1") is not None:
                total_cm1 += float(d["cm1"])
                any_cm1 = True
    parent_doc_id = f"{meeting_week}_{week_key}_{parent_area}" if meeting_week else f"{week_key}_{parent_area}"
    db.collection(FS_PROSPECTS).document(parent_doc_id).set({
        "meeting_week": meeting_week,
        "week_key": week_key,
        "area": parent_area,
        "teu": round(total_teu) if any_teu else None,
        "cm1": round(total_cm1) if any_cm1 else None,
        "updated": datetime.now(tz=JST).isoformat(),
        "auto_sum": True,
    })


def _auto_sum_parent_monthly(parent_area: str, ym: str, meeting_week: str):
    """子エリアの月間見込みを合算して親エリアに保存"""
    children = AREA_AUTO_SUM.get(parent_area, [])
    total_teu = 0.0
    any_teu = False
    for child in children:
        doc_id = f"{meeting_week}_{ym}_{child}" if meeting_week else f"{ym}_{child}"
        doc = db.collection(FS_MONTHLY_PROSPECTS).document(doc_id).get()
        if doc.exists:
            d = doc.to_dict()
            if d.get("teu") is not None:
                total_teu += float(d["teu"])
                any_teu = True
    parent_doc_id = f"{meeting_week}_{ym}_{parent_area}" if meeting_week else f"{ym}_{parent_area}"
    db.collection(FS_MONTHLY_PROSPECTS).document(parent_doc_id).set({
        "meeting_week": meeting_week,
        "ym": ym,
        "area": parent_area,
        "teu": round(total_teu) if any_teu else None,
        "cm1_per_teu": None,
        "updated": datetime.now(tz=JST).isoformat(),
        "auto_sum": True,
    })


@app.route("/api/prospect", methods=["POST"])
@require_editor
def api_save_prospect():
    """見込みデータを Firestore に保存（meeting_week でスコープ）"""
    data = request.json
    week_key = data.get("week_key")
    area = data.get("area")
    meeting_week = data.get("meeting_week", "")
    teu = data.get("teu")
    cm1 = data.get("cm1")

    if not week_key or not area:
        return jsonify({"error": "week_key and area required"}), 400

    # 親エリアへの直接保存をブロック
    if area in AREA_AUTO_SUM:
        return jsonify({"error": f"{area} は自動合算エリアのため直接編集できません"}), 400

    # meeting_week 付きの doc_id で会議週ごとに独立保存
    doc_id = f"{meeting_week}_{week_key}_{area}" if meeting_week else f"{week_key}_{area}"
    db.collection(FS_PROSPECTS).document(doc_id).set({
        "meeting_week": meeting_week,
        "week_key": week_key,
        "area": area,
        "teu": teu,
        "cm1": cm1,
        "updated": datetime.now(tz=JST).isoformat(),
    })

    # ── 子エリア保存時: 親エリアを自動合算 ──
    parent_area = _CHILD_TO_PARENT.get(area)
    if parent_area:
        _auto_sum_parent_prospect(parent_area, week_key, meeting_week)

    return jsonify({"ok": True})


@app.route("/api/note", methods=["GET", "POST"])
@require_login
def api_note():
    """エリアメモを取得/保存"""
    week_key = request.args.get("week_key") or (request.json or {}).get("week_key")
    area = request.args.get("area") or (request.json or {}).get("area")

    if request.method == "GET":
        notes = load_notes(week_key)
        return jsonify(notes)

    # POST
    note = (request.json or {}).get("note", "")
    if _firestore_available:
        doc_id = f"{week_key}_{area}"
        db.collection(FS_NOTES).document(doc_id).set({
            "week_key": week_key,
            "area": area,
            "note": note,
            "updated": datetime.utcnow().isoformat(),
        })
    else:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            INSERT INTO notes (week_key, area, note, updated)
            VALUES (?, ?, ?, datetime('now'))
            ON CONFLICT(week_key, area) DO UPDATE SET
                note=excluded.note, updated=excluded.updated
        """, (week_key, area, note))
        conn.commit()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/image", methods=["GET", "POST", "DELETE"])
@require_login
def api_image():
    """画像アップロード / 取得 / 削除"""
    if request.method == "GET":
        week_key = request.args.get("week_key")
        area = request.args.get("area")
        return jsonify(load_images(week_key, area))

    if request.method == "POST":
        user = get_current_user()
        if not user or user["role"] != "editor":
            return jsonify({"error": "editor role required"}), 403
        week_key = request.form.get("week_key")
        area = request.form.get("area")
        caption = request.form.get("caption", "")
        file = request.files.get("image")
        if not file or not week_key or not area:
            return jsonify({"error": "missing params"}), 400

        import uuid
        ext = Path(file.filename).suffix
        fname = f"{uuid.uuid4().hex}{ext}"

        # GCS にアップロード（利用可能時）、フォールバック: ローカル
        if _gcs_available:
            blob = gcs_bucket.blob(f"meeting-uploads/{fname}")
            blob.upload_from_file(file, content_type=file.content_type)
        else:
            file.save(UPLOAD_DIR / fname)

        # メタデータを Firestore に保存（利用可能時）、フォールバック: SQLite
        if _firestore_available:
            doc_ref = db.collection(FS_IMAGES).document()
            doc_ref.set({
                "week_key": week_key,
                "area": area,
                "filename": fname,
                "caption": caption,
                "uploaded": datetime.utcnow().isoformat(),
            })
            img_id = doc_ref.id
        else:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("""
                INSERT INTO images (week_key, area, filename, caption, uploaded)
                VALUES (?, ?, ?, ?, datetime('now'))
            """, (week_key, area, fname, caption))
            conn.commit()
            img_id = c.lastrowid
            conn.close()
        return jsonify({"id": img_id, "filename": fname, "caption": caption})

    if request.method == "DELETE":
        user = get_current_user()
        if not user or user["role"] != "editor":
            return jsonify({"error": "editor role required"}), 403
        img_id = request.args.get("id")

        if _firestore_available:
            # Firestore からメタデータ取得 → GCS/ローカルのファイル削除 → ドキュメント削除
            doc_ref = db.collection(FS_IMAGES).document(img_id)
            doc = doc_ref.get()
            if doc.exists:
                fname = doc.to_dict().get("filename", "")
                if fname:
                    if _gcs_available:
                        try:
                            gcs_bucket.blob(f"meeting-uploads/{fname}").delete()
                        except Exception:
                            pass
                    else:
                        try:
                            (UPLOAD_DIR / fname).unlink(missing_ok=True)
                        except Exception:
                            pass
                doc_ref.delete()
        else:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            row = c.execute("SELECT filename FROM images WHERE id=?", (img_id,)).fetchone()
            if row:
                try:
                    (UPLOAD_DIR / row[0]).unlink(missing_ok=True)
                except Exception:
                    pass
                c.execute("DELETE FROM images WHERE id=?", (img_id,))
                conn.commit()
            conn.close()
        return jsonify({"ok": True})


@app.route("/api/monthly_prospect", methods=["GET", "POST"])
@require_login
def api_monthly_prospect():
    """月間見込み (TEU / CM1/T) を Firestore で取得・保存"""
    if request.method == "GET":
        ym   = request.args.get("ym")
        area = request.args.get("area")
        doc_id = f"{ym}_{area}"
        doc = db.collection(FS_MONTHLY_PROSPECTS).document(doc_id).get()
        if doc.exists:
            d = doc.to_dict()
            return jsonify({"teu": d.get("teu"), "cm1_per_teu": d.get("cm1_per_teu")})
        return jsonify({"teu": None, "cm1_per_teu": None})

    # POST - editor only
    user = get_current_user()
    if not user or user["role"] != "editor":
        return jsonify({"error": "editor required"}), 403
    data = request.json or {}
    ym   = data.get("ym")
    area = data.get("area")
    meeting_week = data.get("meeting_week", "")
    if not ym or not area:
        return jsonify({"error": "ym and area required"}), 400
    # 親エリアへの直接保存をブロック
    if area in AREA_AUTO_SUM:
        return jsonify({"error": f"{area} は自動合算エリアのため直接編集できません"}), 400
    doc_id = f"{meeting_week}_{ym}_{area}" if meeting_week else f"{ym}_{area}"
    db.collection(FS_MONTHLY_PROSPECTS).document(doc_id).set({
        "meeting_week": meeting_week,
        "ym": ym,
        "area": area,
        "teu": data.get("teu"),
        "cm1_per_teu": data.get("cm1_per_teu"),
        "updated": datetime.now(tz=JST).isoformat(),
    })
    # 子エリア保存時: 親エリアを自動合算
    parent_area = _CHILD_TO_PARENT.get(area)
    if parent_area:
        _auto_sum_parent_monthly(parent_area, ym, meeting_week)
    return jsonify({"ok": True})


# ── ブロックエディター API (Firestore + GCS) ────────────

def _upload_image(file) -> str:
    """画像をGCS（利用可能時）またはローカルに保存し、ファイル名を返す"""
    ext = Path(file.filename).suffix
    fname = f"{uuid.uuid4().hex}{ext}"
    if _gcs_available:
        blob = gcs_bucket.blob(f"meeting-uploads/{fname}")
        blob.upload_from_file(file, content_type=file.content_type)
    else:
        file.save(UPLOAD_DIR / fname)
    return fname


def _delete_image(filename: str):
    """画像をGCSまたはローカルから削除"""
    if not filename:
        return
    if _gcs_available:
        try:
            gcs_bucket.blob(f"meeting-uploads/{filename}").delete()
        except Exception:
            pass
    else:
        try:
            (UPLOAD_DIR / filename).unlink(missing_ok=True)
        except Exception:
            pass


@app.route("/api/blocks", methods=["GET"])
@require_login
def api_blocks_get():
    week_key = request.args.get("week_key")
    area     = request.args.get("area")
    if not week_key or not area:
        return jsonify([])
    try:
        docs = (db.collection(FS_BLOCKS)
                .where("week_key", "==", week_key)
                .where("area", "==", area)
                .stream())
        blocks = []
        for doc in docs:
            d = doc.to_dict()
            blocks.append({
                "id": doc.id,
                "block_type": d.get("block_type"),
                "content": d.get("content", ""),
                "filename": d.get("filename"),
                "img_width": d.get("img_width", 200),
                "block_order": d.get("block_order", 0),
            })
        blocks.sort(key=lambda b: b["block_order"])
        return jsonify(blocks)
    except Exception as e:
        # 複合インデックスが未作成の場合、エラーメッセージにURL含む
        print(f"blocks GET error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/blocks", methods=["POST"])
@require_editor
def api_blocks_post():
    """新規ブロック作成 (JSON: text / multipart: image)"""
    if request.is_json:
        data       = request.json or {}
        block_type = data.get("block_type")
        week_key   = data.get("week_key")
        area       = data.get("area")
        content    = data.get("content", "")
        filename   = None
    else:
        block_type = request.form.get("block_type")
        week_key   = request.form.get("week_key")
        area       = request.form.get("area")
        content    = request.form.get("content", "")
        filename   = None
        if block_type == "image":
            file = request.files.get("image")
            if not file:
                return jsonify({"error": "no image file"}), 400
            filename = _upload_image(file)

    if not week_key or not area or block_type not in ("text", "markdown", "image", "ai"):
        return jsonify({"error": "invalid params"}), 400

    # 最大 block_order を取得 (複合インデックス不要: Python側で算出)
    max_order = -1
    existing = (db.collection(FS_BLOCKS)
                .where("week_key", "==", week_key)
                .where("area", "==", area)
                .stream())
    for doc in existing:
        order = doc.to_dict().get("block_order", -1)
        if order > max_order:
            max_order = order

    new_ref = db.collection(FS_BLOCKS).document()
    new_ref.set({
        "week_key": week_key,
        "area": area,
        "block_order": max_order + 1,
        "block_type": block_type,
        "content": content,
        "filename": filename,
        "img_width": 200,
        "updated": datetime.now(tz=JST).isoformat(),
    })
    return jsonify({"id": new_ref.id, "ok": True})


@app.route("/api/blocks/<block_id>", methods=["PATCH"])
@require_editor
def api_blocks_patch(block_id):
    data = request.json or {}
    update = {"updated": datetime.now(tz=JST).isoformat()}
    if "content" in data:
        update["content"] = data["content"]
    if "img_width" in data:
        update["img_width"] = data["img_width"]
    db.collection(FS_BLOCKS).document(block_id).update(update)
    return jsonify({"ok": True})


@app.route("/api/blocks/<block_id>", methods=["DELETE"])
@require_editor
def api_blocks_delete(block_id):
    doc = db.collection(FS_BLOCKS).document(block_id).get()
    if doc.exists:
        d = doc.to_dict()
        _delete_image(d.get("filename"))
        db.collection(FS_BLOCKS).document(block_id).delete()
    return jsonify({"ok": True})


@app.route("/api/blocks/reorder", methods=["POST"])
@require_editor
def api_blocks_reorder():
    order_list = (request.json or {}).get("order", [])  # [{id, order}, ...]
    batch = db.batch()
    for item in order_list:
        ref = db.collection(FS_BLOCKS).document(str(item["id"]))
        batch.update(ref, {"block_order": item["order"]})
    batch.commit()
    return jsonify({"ok": True})


@app.route("/api/ai-analyze", methods=["POST"])
@require_editor
def api_ai_analyze():
    """AI によるデータ分析。プロンプトとエリアを受け取り、HTML を返す (Claude優先、Geminiフォールバック)"""
    use_claude = bool(_claude_client)
    if not use_claude and not _gemini_client:
        return jsonify({"error": "AI が利用できません (Claude/Gemini 両方未設定)"}), 503

    data = request.json or {}
    prompt = data.get("prompt", "").strip()
    area = data.get("area", "ALL")
    history = data.get("history", [])  # [{role: "user"|"ai", text: "..."}]

    if not prompt:
        return jsonify({"error": "プロンプトが空です"}), 400

    try:
        df = get_bq_df()
        df_area = _filter_area(df, area)

        # エリア別かALLかでデータ提供方法を変える
        raw_cols = ["bq_ym", "week_key", "Booking_Shipper", "POL", "POD", "DLY", "AREA", "TEU", "CM1", "TEU_with_cm1"]
        use_raw = (area != "ALL")  # 個別エリアは生データを渡す

        if use_raw:
            # 生データCSV（AIが自由に集計可能）
            raw_csv = df_area[raw_cols].to_csv(index=False)
            data_section = f"""## 生データ (CSV形式 - 自由に集計してください)
列: bq_ym(年月), week_key(週), Booking_Shipper(荷主), POL(積港), POD(揚港), DLY(最終仕向地), AREA(エリア), TEU, CM1(粗利益), TEU_with_cm1(運賃紐づきTEU)
注意: CM1/TEU = CM1 / TEU_with_cm1 で計算すること（TEU_with_cm1=0の行はCM1/TEU計算から除外）

{raw_csv}"""
        else:
            # ALL: データが大きいため集約サマリーを提供
            monthly = df_area.groupby("bq_ym").agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum"),
                Shipments=("Booking_No_", "nunique"),
            ).reset_index()
            monthly["CM1_per_TEU"] = (monthly["CM1"] / monthly["TEU_with_cm1"].replace(0, float("nan"))).round(0)

            weekly = df_area.groupby("week_key").agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum"),
            ).reset_index()
            weekly["CM1_per_TEU"] = (weekly["CM1"] / weekly["TEU_with_cm1"].replace(0, float("nan"))).round(0)

            shipper = df_area.groupby("Booking_Shipper").agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum"),
            ).reset_index().sort_values("TEU", ascending=False).head(30)
            shipper["CM1_per_TEU"] = (shipper["CM1"] / shipper["TEU_with_cm1"].replace(0, float("nan"))).round(0)

            by_area = df.groupby("AREA").agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum"),
            ).reset_index().sort_values("TEU", ascending=False)
            by_area["CM1_per_TEU"] = (by_area["CM1"] / by_area["TEU_with_cm1"].replace(0, float("nan"))).round(0)

            top_shippers = shipper["Booking_Shipper"].head(20).tolist()
            detail = df_area[df_area["Booking_Shipper"].isin(top_shippers)].groupby(
                ["Booking_Shipper", "bq_ym"]
            ).agg(TEU=("TEU", "sum"), CM1=("CM1", "sum")).reset_index()

            route_summary = df_area.groupby(["POL", "POD"]).agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum"),
            ).reset_index().sort_values("TEU", ascending=False).head(30)
            route_summary["CM1_per_TEU"] = (route_summary["CM1"] / route_summary["TEU_with_cm1"].replace(0, float("nan"))).round(0)

            shipper_route = df_area.groupby(["Booking_Shipper", "POL", "POD", "DLY"]).agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum"),
            ).reset_index().sort_values("TEU", ascending=False).head(50)
            shipper_route["CM1_per_TEU"] = (shipper_route["CM1"] / shipper_route["TEU_with_cm1"].replace(0, float("nan"))).round(0)

            data_section = f"""## 月別サマリー
{monthly.to_csv(index=False)}

## 週別サマリー
{weekly.to_csv(index=False)}

## Shipper別 TEU上位30社
{shipper.to_csv(index=False)}

## エリア別サマリー
{by_area.to_csv(index=False)}

## 上位20社 × 月別明細
{detail.to_csv(index=False)}

## ルート別サマリー (POL→POD) 上位30
{route_summary.to_csv(index=False)}

## Shipper×ルート明細 上位50
{shipper_route.to_csv(index=False)}

## 上位Shipper×ルート×月別明細 (前月比・上昇額分析用、上位20社)
{df_area[df_area["Booking_Shipper"].isin(top_shippers)].groupby(["Booking_Shipper", "POL", "DLY", "bq_ym"]).agg(
    TEU=("TEU", "sum"), CM1=("CM1", "sum"),
    TEU_with_cm1=("TEU_with_cm1", "sum")
).reset_index().assign(
    CM1_per_TEU=lambda x: (x["CM1"] / x["TEU_with_cm1"].replace(0, float("nan"))).round(0)
).drop(columns=["TEU_with_cm1"]).sort_values(["Booking_Shipper", "bq_ym"]).to_csv(index=False)}"""

        # グラフリクエストかどうか判定（現在のプロンプト＋会話履歴も考慮）
        chart_keywords = ["グラフ", "チャート", "chart", "graph"]
        all_texts = prompt.lower() + " " + " ".join(m.get("text", "").lower() for m in history)
        is_chart_request = any(kw in all_texts for kw in chart_keywords)

        if is_chart_request:
            system_prompt = f"""あなたは海運会社 KMTC Japan の Weekly Meeting 用データアナリストです。
提供されるデータに基づき、Chart.js 用のJSONデータを返してください。

## 出力ルール（厳守）
- 必ず有効な JSON のみを返してください。説明文やマークダウンは一切不要です。
- ```json タグも不要。純粋な JSON のみ。
- 以下の形式で返してください:
{{
  "title": "グラフのタイトル",
  "type": "bar" または "line" または "bar+line",
  "labels": ["ラベル1", "ラベル2", ...],
  "datasets": [
    {{
      "label": "データセット名",
      "data": [数値1, 数値2, ...],
      "type": "bar" または "line" (bar+lineの場合),
      "backgroundColor": "#色コード" (barの場合),
      "borderColor": "#色コード" (lineの場合),
      "yAxisID": "y" (左軸) または "y1" (右軸)
    }}
  ],
  "y_label": "左軸ラベル",
  "y1_label": "右軸ラベル (あれば)"
}}
- 棒グラフの色: "#3f51b5" (青), "#ff9800" (オレンジ), "#4caf50" (緑), "#e91e63" (ピンク), "#9c27b0" (紫)
- 折れ線グラフの色: "#c62828" (赤), "#2e7d32" (緑), "#1565c0" (青)
- TEUは左軸(y), CM1/TEUは右軸(y1)
- 数値は整数
- 生データが提供されている場合は、自分で集計してからグラフデータを作成してください

## データ
対象エリア: {area}

{data_section}
"""
        else:
            system_prompt = f"""あなたは海運会社 KMTC Japan の Weekly Meeting 用データアナリストです。
提供されるデータに基づき、正確で簡潔な分析をHTMLで返してください。

## 出力ルール
- HTML のみ返してください（```html タグ不要、<html><body>も不要）
- <table>, <ul>, <ol>, <p>, <strong> 等を使った見やすいフォーマット
- テーブルには必ず style="border-collapse:collapse; width:100%" を付与
- th には style="background:#1a237e; color:#fff; padding:6px 10px; text-align:center; font-size:13px"
- td には style="border:1px solid #ddd; padding:5px 8px; text-align:right; font-size:13px"
- 左揃えにすべき列（Shipper名等）は text-align:left
- 数値はカンマ区切り (例: 1,234)
- CM1/TEU は整数の$表記 (例: $342)
- 説明文は簡潔に日本語で
- 生データが提供されている場合は、自分で集計してから分析結果を作成してください

## データ
対象エリア: {area}

{data_section}
"""

        if use_claude:
            # ── Claude API ──
            messages = []
            for msg in history:
                role = "user" if msg.get("role") == "user" else "assistant"
                messages.append({"role": role, "content": msg.get("text", "")})
            messages.append({"role": "user", "content": prompt})

            response = _claude_client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                system=system_prompt,
                messages=messages,
                temperature=0.3,
            )
            result_text = response.content[0].text.strip()
        else:
            # ── Gemini API (フォールバック) ──
            contents = []
            for msg in history:
                role = "user" if msg.get("role") == "user" else "model"
                contents.append(genai.types.Content(
                    role=role,
                    parts=[genai.types.Part.from_text(text=msg.get("text", ""))]
                ))
            contents.append(genai.types.Content(
                role="user",
                parts=[genai.types.Part.from_text(text=prompt)]
            ))

            response = _gemini_client.models.generate_content(
                model="gemini-2.0-flash",
                contents=contents,
                config=genai.types.GenerateContentConfig(
                    system_instruction=system_prompt,
                    temperature=0.3,
                    max_output_tokens=4096,
                ),
            )
            result_text = (response.text or "").strip()
        # コードブロックのマーカーを除去
        if result_text.startswith("```json"):
            result_text = result_text[7:]
        if result_text.startswith("```html"):
            result_text = result_text[7:]
        if result_text.startswith("```"):
            result_text = result_text[3:]
        if result_text.endswith("```"):
            result_text = result_text[:-3]
        result_text = result_text.strip()

        # レスポンスの自動判定: JSONならチャート、それ以外はHTML
        if result_text.startswith("{"):
            try:
                chart_data = json.loads(result_text)
                # chart データに必要なキーがあるか確認
                if "datasets" in chart_data or "labels" in chart_data:
                    return jsonify({"ok": True, "chart": chart_data})
            except json.JSONDecodeError:
                pass
        return jsonify({"ok": True, "html": result_text})

    except Exception as e:
        print(f"AI analyze error: {e}")
        return jsonify({"error": f"AI分析エラー: {str(e)}"}), 500


@app.route("/api/ai-template-comment", methods=["POST"])
@require_editor
def api_ai_template_comment():
    """Claude AI: テンプレートブロック専用のAIコメント生成"""
    if not _claude_client:
        return jsonify({"error": "Claude AI が利用できません (ANTHROPIC_API_KEY 未設定)"}), 503

    data = request.json or {}
    template_id = data.get("template_id", "")
    area = data.get("area", "ALL")
    meeting_week = data.get("meeting_week", "")
    lang = data.get("lang", "ja")  # ja / en / kr

    if not template_id:
        return jsonify({"error": "template_id が指定されていません"}), 400

    try:
        today = date.today()
        df = get_bq_df()
        tpl_data = build_template_data(area, df, today, meeting_week)

        # テンプレートのラベル取得
        tpl_def = next((d for d in TEMPLATE_DEFINITIONS if d["id"] == template_id), None)
        block_label = tpl_def["label"] if tpl_def else template_id

        # テンプレート固有データ抽出
        block_data = tpl_data.get(template_id)

        # monthly/weekly 系は追加の文脈データを付与
        context_data = {}
        if template_id in ("booking_monthly", "booking_weekly"):
            context_data["booking_count"] = tpl_data.get("booking_count", {})
        elif template_id.startswith("shipper_"):
            context_data[template_id] = block_data
            # 比較用に当月・翌月両方の増減も参考情報として提供
            for k in ["shipper_increase_curr", "shipper_increase_next",
                       "shipper_decrease_curr", "shipper_decrease_next"]:
                if k != template_id and k in tpl_data:
                    context_data[k + "_ref"] = tpl_data[k]
        elif template_id.startswith("combo_"):
            context_data[template_id] = block_data
        else:
            context_data[template_id] = block_data

        # ================================================================
        # 共通文脈データ (全テンプレートで使えるように)
        # ================================================================
        df_area = _filter_area(df, area)
        ref = _ref_date_from_meeting_week(meeting_week, today)
        months_6 = _cal_months(ref.year, ref.month, 5, 0)
        yms_6 = [f"{y}-{m:02d}" for y, m in months_6]
        prev_ym = yms_6[-2] if len(yms_6) >= 2 else None
        curr_ym_ctx = yms_6[-1]
        df_2m = df_area[df_area["bq_ym"].isin(yms_6[-2:])]

        # (A) 月間推移 6か月
        ms = df_area[df_area["bq_ym"].isin(yms_6)].groupby("bq_ym").agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum"),
            TEU_with_cm1=("TEU_with_cm1", "sum")
        ).reset_index()
        ms["CM1_per_TEU"] = (ms["CM1"] / ms["TEU_with_cm1"].replace(0, float("nan"))).round(0)
        context_data["_monthly_trend"] = ms.drop(columns=["TEU_with_cm1"]).to_dict(orient="records")

        # (B) 仕向地 (DLY) 別 内訳 — TEU + CM1/TEU
        ctr_2m = df_2m.groupby(["bq_ym", "DLY"]).agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum"),
            TEU_with_cm1=("TEU_with_cm1", "sum")
        ).reset_index()
        ctr_2m["CM1_per_TEU"] = (ctr_2m["CM1"] / ctr_2m["TEU_with_cm1"].replace(0, float("nan"))).round(0)
        ctr_2m = ctr_2m.drop(columns=["TEU_with_cm1"]).sort_values(["bq_ym", "TEU"], ascending=[True, False])
        context_data["_dly_breakdown_top10"] = ctr_2m.groupby("bq_ym").head(10).to_dict(orient="records")

        # (C) 荷主別内訳 Top15 (直近2か月) — TEU + CM1/TEU
        shipper_2m = df_2m.groupby(["bq_ym", "Booking_Shipper"]).agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum"),
            TEU_with_cm1=("TEU_with_cm1", "sum")
        ).reset_index()
        shipper_2m["CM1_per_TEU"] = (shipper_2m["CM1"] / shipper_2m["TEU_with_cm1"].replace(0, float("nan"))).round(0)
        shipper_2m = shipper_2m.drop(columns=["TEU_with_cm1"]).sort_values(["bq_ym", "TEU"], ascending=[True, False])
        context_data["_shipper_top15_2months"] = shipper_2m.groupby("bq_ym").head(15).to_dict(orient="records")

        # (D) 荷主×航路 (POL→DLY) 詳細 Top20 (直近2か月) — TEU + CM1/TEU
        shipper_route_2m = df_2m.groupby(["bq_ym", "Booking_Shipper", "POL", "DLY"]).agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum"),
            TEU_with_cm1=("TEU_with_cm1", "sum")
        ).reset_index()
        shipper_route_2m["CM1_per_TEU"] = (shipper_route_2m["CM1"] / shipper_route_2m["TEU_with_cm1"].replace(0, float("nan"))).round(0)
        shipper_route_2m = shipper_route_2m.drop(columns=["TEU_with_cm1"]).sort_values(["bq_ym", "TEU"], ascending=[True, False])
        context_data["_shipper_route_top20"] = shipper_route_2m.groupby("bq_ym").head(20).to_dict(orient="records")

        # (E) 営業マン (POL_Sales) 別 内訳 — TEU + CM1/TEU + 担当荷主
        if "POL_Sales" in df_area.columns:
            sales_2m = df_2m.groupby(["bq_ym", "POL_Sales"]).agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum")
            ).reset_index()
            sales_2m["CM1_per_TEU"] = (sales_2m["CM1"] / sales_2m["TEU_with_cm1"].replace(0, float("nan"))).round(0)
            sales_2m = sales_2m.drop(columns=["TEU_with_cm1"]).sort_values(["bq_ym", "TEU"], ascending=[True, False])
            context_data["_sales_top10"] = sales_2m.groupby("bq_ym").head(10).to_dict(orient="records")

            # 営業マン別の主要荷主 (Top3 per salesperson, 直近月)
            sales_shipper = df_area[df_area["bq_ym"] == curr_ym_ctx].groupby(
                ["POL_Sales", "Booking_Shipper"]
            ).agg(TEU=("TEU", "sum")).reset_index().sort_values("TEU", ascending=False)
            context_data["_sales_shipper_top3"] = sales_shipper.groupby("POL_Sales").head(3).to_dict(orient="records")

            # 荷主→主担当営業マン マッピング (TEU最大の営業を主担当とする)
            shipper_primary = sales_shipper.sort_values("TEU", ascending=False).drop_duplicates(
                subset=["Booking_Shipper"], keep="first"
            )[["Booking_Shipper", "POL_Sales", "TEU"]].head(20)
            context_data["_shipper_primary_sales"] = shipper_primary.to_dict(orient="records")

        # (F) POL (積地) 別 内訳 — TEU + CM1/TEU
        pol_2m = df_2m.groupby(["bq_ym", "POL"]).agg(
            TEU=("TEU", "sum"), CM1=("CM1", "sum"),
            TEU_with_cm1=("TEU_with_cm1", "sum")
        ).reset_index()
        pol_2m["CM1_per_TEU"] = (pol_2m["CM1"] / pol_2m["TEU_with_cm1"].replace(0, float("nan"))).round(0)
        pol_2m = pol_2m.drop(columns=["TEU_with_cm1"]).sort_values(["bq_ym", "TEU"], ascending=[True, False])
        context_data["_pol_breakdown_top10"] = pol_2m.groupby("bq_ym").head(10).to_dict(orient="records")

        # (F2) POL Count用: 前月→今月のPOL出現/消失 + POL×荷主内訳
        if template_id == "pol_count" and prev_ym:
            prev_pols = set(pol_2m[pol_2m["bq_ym"] == prev_ym]["POL"].tolist())
            curr_pols = set(pol_2m[pol_2m["bq_ym"] == curr_ym_ctx]["POL"].tolist())
            context_data["_pol_new"] = sorted(curr_pols - prev_pols)        # 今月新たに出現
            context_data["_pol_lost"] = sorted(prev_pols - curr_pols)       # 今月消失
            # POL×荷主 (今月の各POLの主要荷主Top3)
            pol_shipper = df_area[df_area["bq_ym"] == curr_ym_ctx].groupby(
                ["POL", "Booking_Shipper"]
            ).agg(TEU=("TEU", "sum")).reset_index().sort_values("TEU", ascending=False)
            context_data["_pol_shipper_top3"] = pol_shipper.groupby("POL").head(3).to_dict(orient="records")
            # 前月のPOL×荷主 (比較用)
            if prev_ym:
                pol_shipper_prev = df_area[df_area["bq_ym"] == prev_ym].groupby(
                    ["POL", "Booking_Shipper"]
                ).agg(TEU=("TEU", "sum")).reset_index().sort_values("TEU", ascending=False)
                context_data["_pol_shipper_prev_top3"] = pol_shipper_prev.groupby("POL").head(3).to_dict(orient="records")

        # ================================================================
        # テンプレート固有の追加文脈データ
        # ================================================================

        # Combo系: 荷主×航路別の CM1/TEU 変動 Top5
        if template_id.startswith("combo_") and prev_ym:
            combo_cm1 = df_2m.groupby(
                ["bq_ym", "Booking_Shipper", "POL", "DLY"]
            ).agg(TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                  TEU_with_cm1=("TEU_with_cm1", "sum")).reset_index()
            combo_cm1["CM1_per_TEU"] = (combo_cm1["CM1"] / combo_cm1["TEU_with_cm1"].replace(0, float("nan"))).round(0)
            combo_cm1 = combo_cm1[combo_cm1["CM1_per_TEU"].notna() & (combo_cm1["CM1_per_TEU"] != 0)]
            combo_cm1 = combo_cm1.drop(columns=["TEU_with_cm1"])
            piv = combo_cm1.pivot_table(
                index=["Booking_Shipper", "POL", "DLY"],
                columns="bq_ym", values="CM1_per_TEU", aggfunc="first"
            )
            if prev_ym in piv.columns and curr_ym_ctx in piv.columns:
                piv = piv.dropna(subset=[prev_ym, curr_ym_ctx])
                piv["cm1t_change"] = piv[curr_ym_ctx] - piv[prev_ym]
                def _cm1t_rows(sub):
                    return [{"shipper": idx[0], "pol": idx[1], "dly": idx[2],
                             f"cm1t_{prev_ym}": round(float(r[prev_ym])),
                             f"cm1t_{curr_ym_ctx}": round(float(r[curr_ym_ctx])),
                             "change": round(float(r["cm1t_change"]))}
                            for idx, r in sub.iterrows()]
                context_data["_cm1t_combo_increase"] = _cm1t_rows(piv.nlargest(5, "cm1t_change"))
                context_data["_cm1t_combo_decrease"] = _cm1t_rows(piv.nsmallest(5, "cm1t_change"))

        # Trade Lane用: 仕向地ごとの主要荷主 Top3
        if template_id == "trade_lane" and prev_ym:
            tl_shipper = df_2m.groupby(["bq_ym", "DLY", "Booking_Shipper"]).agg(
                TEU=("TEU", "sum")
            ).reset_index().sort_values(["bq_ym", "DLY", "TEU"], ascending=[True, True, False])
            context_data["_trade_lane_shippers"] = tl_shipper.groupby(["bq_ym", "DLY"]).head(3).to_dict(orient="records")

        # CM1 Range用: 荷主のセグメント移動 (High/Mid/Low間の推移)
        if template_id == "cm1_range":
            cm1_range_data = tpl_data.get("cm1_range", {})
            for ym in yms_6[-2:]:
                sub = df_area[df_area["bq_ym"] == ym]
                sa = sub.groupby("Booking_Shipper").agg(
                    TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                    TEU_with_cm1=("TEU_with_cm1", "sum")
                ).reset_index()
                sa = sa[sa["TEU"] > 0]
                sa["CM1_per_TEU"] = (sa["CM1"] / sa["TEU_with_cm1"].replace(0, float("nan"))).round(0)
                sa = sa.drop(columns=["TEU_with_cm1"])
                q_data = cm1_range_data.get(ym, {})
                q75 = q_data.get("q75", 0)
                q25 = q_data.get("q25", 0)
                sa["segment"] = sa["CM1_per_TEU"].apply(
                    lambda x: "High" if x >= q75 else ("Low" if x < q25 else "Mid"))
                context_data[f"_cm1_shipper_{ym}"] = sa.sort_values(
                    "TEU", ascending=False).head(15).to_dict(orient="records")
            # セグメント移動した荷主を特定
            if prev_ym and f"_cm1_shipper_{prev_ym}" in context_data and f"_cm1_shipper_{curr_ym_ctx}" in context_data:
                prev_segs = {r["Booking_Shipper"]: r["segment"] for r in context_data[f"_cm1_shipper_{prev_ym}"]}
                curr_segs = {r["Booking_Shipper"]: r["segment"] for r in context_data[f"_cm1_shipper_{curr_ym_ctx}"]}
                movers = []
                for s in set(prev_segs) & set(curr_segs):
                    if prev_segs[s] != curr_segs[s]:
                        movers.append({"shipper": s, "from": prev_segs[s], "to": curr_segs[s]})
                if movers:
                    context_data["_cm1_segment_movers"] = movers

        # Sales Contribution用: 営業マン別の荷主構成変化
        if template_id == "sales_contribution" and prev_ym:
            # 荷主→主担当マッピング (TEU最大の営業を主担当とする, 全期間)
            _all_sales_shipper = df_area.groupby(
                ["POL_Sales", "Booking_Shipper"]
            ).agg(TEU=("TEU", "sum")).reset_index().sort_values("TEU", ascending=False)
            _primary_map = _all_sales_shipper.drop_duplicates(
                subset=["Booking_Shipper"], keep="first"
            ).set_index("Booking_Shipper")["POL_Sales"].to_dict()

            # 営業マン別シェア (前月 vs 当月)
            for ym in [prev_ym, curr_ym_ctx]:
                sub = df_area[df_area["bq_ym"] == ym]
                ss = sub.groupby(["POL_Sales", "Booking_Shipper"]).agg(
                    TEU=("TEU", "sum")
                ).reset_index().sort_values("TEU", ascending=False)
                # 主担当フラグ追加: この営業マンがTEU最大の担当者か
                ss["is_primary"] = ss.apply(
                    lambda r: _primary_map.get(r["Booking_Shipper"]) == r["POL_Sales"], axis=1)
                context_data[f"_sales_detail_{ym}"] = ss.groupby("POL_Sales").head(3).to_dict(orient="records")

        # New/Regain Customer用: 荷主の航路詳細
        if template_id in ("new_customer", "regain_customer"):
            cust_data = block_data or {}
            cust_list = [c["shipper"] for c in cust_data.get("customers", [])]
            if cust_list:
                cust_routes = df_area[
                    (df_area["bq_ym"] == curr_ym_ctx) &
                    (df_area["Booking_Shipper"].isin(cust_list))
                ].groupby(["Booking_Shipper", "POL", "DLY"]).agg(
                    TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                    TEU_with_cm1=("TEU_with_cm1", "sum")
                ).reset_index()
                cust_routes["CM1_per_TEU"] = (cust_routes["CM1"] / cust_routes["TEU_with_cm1"].replace(0, float("nan"))).round(0)
                cust_routes = cust_routes.drop(columns=["TEU_with_cm1"]).sort_values("TEU", ascending=False)
                context_data["_customer_routes"] = cust_routes.head(20).to_dict(orient="records")

        # CM1 Waterfall用: 要因分解の主要因荷主
        if template_id == "cm1_waterfall" and prev_ym:
            # Rate Effect の主要因: CM1/TEU が大きく変化した荷主
            wf_shipper = df_2m.groupby(["bq_ym", "Booking_Shipper"]).agg(
                TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                TEU_with_cm1=("TEU_with_cm1", "sum")
            ).reset_index()
            wf_shipper["CM1_per_TEU"] = (wf_shipper["CM1"] / wf_shipper["TEU_with_cm1"].replace(0, float("nan"))).round(0)
            wf_piv = wf_shipper.pivot_table(
                index="Booking_Shipper", columns="bq_ym",
                values=["TEU", "CM1_per_TEU"], aggfunc="first"
            )
            if ("CM1_per_TEU", prev_ym) in wf_piv.columns and ("CM1_per_TEU", curr_ym_ctx) in wf_piv.columns:
                wf_piv_flat = pd.DataFrame({
                    "prev_teu": wf_piv[("TEU", prev_ym)],
                    "curr_teu": wf_piv[("TEU", curr_ym_ctx)],
                    "prev_cm1t": wf_piv[("CM1_per_TEU", prev_ym)],
                    "curr_cm1t": wf_piv[("CM1_per_TEU", curr_ym_ctx)],
                }).dropna()
                wf_piv_flat["rate_chg"] = wf_piv_flat["curr_cm1t"] - wf_piv_flat["prev_cm1t"]
                wf_piv_flat["vol_chg"] = wf_piv_flat["curr_teu"] - wf_piv_flat["prev_teu"]
                context_data["_wf_rate_top5"] = wf_piv_flat.nlargest(5, "rate_chg").reset_index().rename(
                    columns={"index": "shipper"}).to_dict(orient="records")
                context_data["_wf_rate_bottom5"] = wf_piv_flat.nsmallest(5, "rate_chg").reset_index().rename(
                    columns={"index": "shipper"}).to_dict(orient="records")
                context_data["_wf_vol_top5"] = wf_piv_flat.nlargest(5, "vol_chg").reset_index().rename(
                    columns={"index": "shipper"}).to_dict(orient="records")

        # Koshi (古紙) 荷主用: 航路別詳細
        if template_id == "koshi_shipper":
            koshi_data = block_data or {}
            all_koshi = set()
            for it in koshi_data.get("items", []):
                all_koshi.add(it.get("shipper", ""))
            if all_koshi:
                kr = df_2m[df_2m["Booking_Shipper"].isin(all_koshi)].groupby(
                    ["bq_ym", "Booking_Shipper", "POL", "DLY"]
                ).agg(TEU=("TEU", "sum"), CM1=("CM1", "sum"),
                      TEU_with_cm1=("TEU_with_cm1", "sum")).reset_index()
                kr["CM1_per_TEU"] = (kr["CM1"] / kr["TEU_with_cm1"].replace(0, float("nan"))).round(0)
                kr = kr.drop(columns=["TEU_with_cm1"]).sort_values(["bq_ym", "TEU"], ascending=[True, False])
                context_data["_koshi_routes"] = kr.groupby(["bq_ym", "Booking_Shipper"]).head(3).to_dict(orient="records")

        # ── 当月データ進捗情報を付与 ──
        curr_ym = f"{ref.year}-{ref.month:02d}"
        import calendar
        days_in_month = calendar.monthrange(ref.year, ref.month)[1]
        elapsed_days = min(today.day, days_in_month) if f"{today.year}-{today.month:02d}" == curr_ym else days_in_month
        month_progress_pct = round(elapsed_days / days_in_month * 100)
        if month_progress_pct < 100:
            # 当月のCM1有無率を計算
            df_curr = df_area[df_area["bq_ym"] == curr_ym]
            total_records = len(df_curr)
            cm1_nonzero = len(df_curr[df_curr["CM1"] != 0]) if total_records > 0 else 0
            cm1_coverage = round(cm1_nonzero / total_records * 100) if total_records > 0 else 0
            context_data["_data_note"] = (
                f"⚠ {curr_ym} is {month_progress_pct}% elapsed ({elapsed_days}/{days_in_month} days). "
                f"CM1 data coverage: {cm1_coverage}% ({cm1_nonzero}/{total_records} records have CM1≠0). "
                f"Many shippers may show CM1/TEU=$0 because freight data has not yet been finalized. "
                f"Do NOT flag CM1=0 as anomaly — it is expected for incomplete months."
            )

        block_json = json.dumps(context_data, ensure_ascii=False, default=str)
        # トークン節約: 大きすぎる場合は切り詰め
        if len(block_json) > 20000:
            block_json = block_json[:20000] + "\n... (truncated)"

        lang_instruction = {
            "ja": "日本語で記載",
            "en": "Write in English",
            "kr": "한국어로 작성",
        }.get(lang, "日本語で記載")

        # テンプレート固有の追加指示
        template_specific = ""
        if template_id == "cm1_range":
            template_specific = """
## CM1 Range 固有ルール
- CM1合計金額の話は不要。CM1/TEU単価の変化に集中する
- Good: CM1/TEU上位(High)の荷主名とTEU、CM1/TEU単価を具体的に記載
- Bad: CM1/TEU下位(Low)の荷主名とTEU、CM1/TEU単価を記載
- _cm1_segment_movers があれば、セグメント移動した荷主を必ず記載（例: SHIPPER_X: Mid→High）
- Q75/Q25の閾値変化も記載（例: Q75: $500→$450）"""
        elif template_id in ("shipper_increase_curr", "shipper_increase_next"):
            template_specific = """
## 増加荷主 固有ルール
- 【Bad】セクションは不要。出力しないこと。
- 表に既に記載されているTop3荷主の主要因航路（Remark列）の繰り返しは不要
- 担当営業マン名の記載は不要
- 【Good】として、以下の別視点で分析:
  - 新規案件の有無: 過去3か月に実績ゼロだった荷主 or 航路（POL→DLY）が今月出現しているか (_shipper_top15_2months, _shipper_route_top20 参照)
  - 増加荷主の社数と全体に占めるインパクト
  - CM1/TEU単価の傾向 (_dly_breakdown_top10 参照): 増加案件が高単価か低単価か"""
        elif template_id in ("shipper_decrease_curr", "shipper_decrease_next"):
            template_specific = """
## 減少荷主 固有ルール
- 【Good】セクションは不要。出力しないこと。
- 表に既に記載されているTop3荷主の主要因航路（Remark列）の繰り返しは不要
- 担当営業マン名の記載は不要
- 【Bad】として、以下の別視点で分析:
  - 喪失案件の有無: 過去3か月に実績があったのに今月ゼロになった荷主 or 航路（POL→DLY） (_shipper_top15_2months, _shipper_route_top20 参照)
  - 減少荷主の社数と全体に占めるインパクト
  - CM1/TEU単価の傾向: 減少した案件が高単価だったか低単価だったか (_dly_breakdown_top10 参照)"""
        elif template_id.startswith("combo_"):
            template_specific = """
## 荷主×航路 固有ルール
- 各行の TEU変化 と CM1/TEU変化 の両面でコメント
- _cm1t_combo_increase/decrease から CM1/TEU変動の大きい荷主×航路も言及
- CM1/TEU=$0 の行はCM1/TEU比較から除外し、TEU変化のみ言及"""
        elif template_id == "trade_lane":
            template_specific = """
## Trade Lane 固有ルール
- 各国/仕向地のTEU変化に加え、_trade_lane_shippers から主要因Shipperを記載
- 例: 「【PH】SHIPPER_A +80T, SHIPPER_B +45T」
- CM1/TEU変化も _dly_breakdown_top10 から確認して記載"""
        elif template_id == "sales_contribution":
            template_specific = """
## Sales Contribution 固有ルール
- シェア変動の大きい営業マンを記載（例: TAOKADA 20%→22%）
- _sales_detail から各営業マンの主要荷主名を記載
- _sales_top10 から CM1/TEU変化も確認
- 重要: _sales_detail の各荷主には is_primary フラグあり。is_primary=False の荷主はその営業マンの「主担当荷主」ではない（別の営業マンが主担当）。
  - 営業マンの実績説明で is_primary=False の荷主を「担当荷主」として記載してはいけない
  - is_primary=True の荷主のみを「担当荷主」として記載すること
- _shipper_primary_sales が荷主→主担当営業マンの正式マッピング。これに従うこと"""
        elif template_id == "cm1_waterfall":
            template_specific = """
## CM1 Waterfall 固有ルール
- Rate Effect の主要因: _wf_rate_top5/_wf_rate_bottom5 からCM1/TEU変化の大きい荷主を記載
- Volume Effect の主要因: _wf_vol_top5 からTEU変化の大きい荷主を記載
- 各荷主名と具体的な変化量を記載"""
        elif template_id == "koshi_shipper":
            template_specific = """
## 古紙荷主 固有ルール
- 過去3か月平均TEU vs 対象月のGapに注目し、増減の大きい荷主を重点コメント
- _koshi_routes から各荷主の主要航路（POL→DLY）とTEU変化を記載
- CM1/TEU変化も航路別に記載"""
        elif template_id in ("new_customer", "regain_customer"):
            template_specific = """
## New/Regain Customer 固有ルール
- _customer_routes から各顧客の航路（POL→DLY）とTEU、CM1/TEUを記載
- 注目すべき大口顧客やCM1/TEUの高い顧客を記載"""
        elif template_id in ("booking_monthly", "booking_weekly"):
            template_specific = """
## Booking件数 固有ルール
- 件数の増減要因を _shipper_top15_2months からShipper名で特定
- Main件数 vs Local件数の構成変化があれば記載"""
        elif template_id == "pol_count":
            template_specific = """
## POL数 固有ルール（POL Count = 出港地のバリエーション数）
- POL数の増減理由を具体的に説明すること
- _pol_new: 今月新たに出現したPOL → どの荷主の案件で増えたか _pol_shipper_top3 から特定
- _pol_lost: 今月消失したPOL → 前月はどの荷主がいたか _pol_shipper_prev_top3 から特定
- TEU変動が大きいPOLは _pol_breakdown_top10 で前月比を確認し、_pol_shipper_top3 から主要荷主名を記載
- 「新規POL: XXX (SHIPPER_A 50T)」「消失POL: YYY (前月 SHIPPER_B 30T)」のように港名+荷主+数量で記載"""

        system_prompt = f"""海運会社 KMTC Japan の会議資料用。公式資料のため、数値の事実のみ記載。

## 絶対禁止（違反すると資料として使えない）
- 理由・原因の推測（「撤退傾向」「需要減」「好調」「期待」等）
- 将来の予測や提案
- 「〜と考えられる」「〜の影響」「〜が寄与」等の因果推論
- 「データ異常」「データ未反映」等のデータ品質コメント

## 当月データに関する重要注意
- _data_note がある場合、当月はまだ月途中
- CM1/TEU はすでに「運賃(O_F)が入っているデータのみ」で計算済み（TEU_with_cm1方式）
- よって CM1/TEU=$0 はその荷主×航路に運賃データが全く無いことを意味する
- CM1/TEU=$0 の行は CM1/TEU比較から除外し、TEU変化のみ言及すること

## フォーマット（厳守）
- 良い点は <span class="ai-good">【Good】</span> で始まるセクション、悪い点は <span class="ai-bad">【Bad】</span> で始まるセクションに分ける
- 各セクションは <span class="ai-good">【Good】</span> または <span class="ai-bad">【Bad】</span> の後に <ul><li> で箇条書き
- 表に既に表示されている数値（合計TEU、CM1等）の記載は不要
- 表の数値の「背景」＝主要因となったShipper名と変化量を記載する
- 国は2文字コード（KR, PH等）、CM1は$表記、TEUはTEU
- 各セクション1〜3行。簡潔に
- {lang_instruction}
{template_specific}
## 良い例
<span class="ai-good">【Good】</span>
<ul><li>PH増加の主因: SHIPPER_A +300T, SHIPPER_B +150T</li></ul>
<span class="ai-bad">【Bad】</span>
<ul><li>AE減少の主因: SHIPPER_C 180→20T</li></ul>

## 悪い例（禁止）
- 「PH: 871→1,324 TEU」→ 表を見ればわかる。Shipper名を書く
- 「撤退傾向」→ 推測禁止
- 「好調」「安定」→ 評価禁止
- 「CM1が0でデータ異常」→ CM1/TEUは運賃確定分のみで計算済。$0は除外して言及しない

## データ
エリア: {area} / 基準日: {ref.strftime('%Y-%m-%d')}

{block_json}"""

        response = _claude_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=500,
            system=system_prompt,
            messages=[{"role": "user", "content": "<span class=\"ai-good\">【Good】</span> と <span class=\"ai-bad\">【Bad】</span> に分けて箇条書き。表の数値の繰り返しは不要、その背景（主要因Shipper名と変化量）だけ記載。推測禁止。"}],
            temperature=0.2,
        )

        result_text = response.content[0].text.strip()
        # コードブロックのマーカーを除去
        for prefix in ["```html", "```"]:
            if result_text.startswith(prefix):
                result_text = result_text[len(prefix):]
        if result_text.endswith("```"):
            result_text = result_text[:-3]
        result_text = result_text.strip()

        return jsonify({"ok": True, "html": result_text})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"AIコメントエラー: {str(e)}"}), 500


@app.route("/api/refresh-status")
@require_login
def api_refresh_status():
    """最終BQ更新日時を返す"""
    info = get_last_refresh()
    if info:
        return jsonify({
            "last_refresh": info.get("last_refresh"),
            "refreshed_by": info.get("refreshed_by", ""),
        })
    return jsonify({"last_refresh": None, "refreshed_by": ""})


@app.route("/api/refresh", methods=["POST"])
@require_editor
def api_refresh():
    """BQデータを手動更新する（編集者のみ）"""
    user = get_current_user()
    ok = do_refresh_bq(update_firestore=True, refreshed_by=user["email"])
    if ok:
        info = get_last_refresh()
        return jsonify({
            "ok": True,
            "last_refresh": info.get("last_refresh") if info else None,
            "refreshed_by": user["email"],
        })
    else:
        return jsonify({"error": "BigQuery更新に失敗しました"}), 500


@app.route("/api/debug-routes")
def debug_routes():
    routes = [str(r) for r in app.url_map.iter_rules()]
    return jsonify(sorted(routes))

# ── 木曜自動アーカイブ (毎週木曜10:00 JST) ──────────────
def auto_archive_current_week():
    """今週のアーカイブを自動保存"""
    with app.app_context():
        try:
            today = date.today()
            week_info = get_week_info(today)
            week_key = week_info["week_key"]

            # BQ データ再取得
            ok = do_refresh_bq(update_firestore=True, refreshed_by="auto-archive")
            if not ok:
                print(f"Auto-archive: BQ refresh failed for {week_key}")
                return

            df = get_bq_df()
            areas = get_all_areas(df)
            areas_data = {}
            for a in areas:
                areas_data[a] = build_summary_for_area(a, df, today, meeting_week=week_key)

            now_jst = datetime.now(tz=JST).isoformat()
            snapshot_doc = {
                "week_key": week_key,
                "created_at": now_jst,
                "created_by": "auto-archive (木曜自動)",
                "areas": areas_data,
            }
            db.collection(FS_SNAPSHOTS).document(week_key).set(snapshot_doc)
            print(f"Auto-archive: {week_key} saved at {now_jst}")
        except Exception as e:
            print(f"Auto-archive error: {e}")


# ── 週マッピング管理 API ─────────────────────────────────
@app.route("/api/week-mapping/sync", methods=["POST"])
def api_week_mapping_sync():
    """
    GAS Web App 経由でスプレッドシートから週マッピングを更新。
    POST body: {"gas_url": "..."} or empty (環境変数/設定値を使用)
    """
    import week_mapping as _wm
    body = request.get_json(silent=True) or {}
    gas_url = body.get("gas_url") or _wm.GAS_WEEK_URL or os.environ.get("GAS_WEEK_URL", "")
    if not gas_url:
        return jsonify(ok=False, error="GAS_WEEK_URL 未設定"), 400
    ok = _refresh_week_from_gas(db=db, gas_url=gas_url)
    if ok:
        # GAS URL を永続化 (次回起動時にも使えるように)
        if db and gas_url != _wm.GAS_WEEK_URL:
            _wm.GAS_WEEK_URL = gas_url
            try:
                db.collection("meeting_config").document("week_mapping_config").set(
                    {"gas_url": gas_url}, merge=True
                )
            except Exception:
                pass
        return jsonify(ok=True, message="週マッピング更新完了",
                       month_week_map={str(y): {str(m): wks for m, wks in ms.items()}
                                       for y, ms in MONTH_WEEK_MAP.items()})
    return jsonify(ok=False, error="GAS からの取得に失敗"), 500


@app.route("/api/week-mapping/upload", methods=["POST"])
def api_week_mapping_upload():
    """
    ブラウザから直接 CSV データを受け取って週マッピングを更新。
    POST body: {"rows": [{"date":"2025-06-29","year":25,"week":27}, ...]}
    """
    body = request.get_json(silent=True) or {}
    rows = body.get("rows", [])
    if not rows:
        return jsonify(ok=False, error="rows が空です"), 400
    ok = _apply_week_rows(rows)
    if ok and db:
        _save_week_to_fs(db, rows)
    if ok:
        return jsonify(ok=True, message="週マッピング更新完了",
                       month_week_map={str(y): {str(m): wks for m, wks in ms.items()}
                                       for y, ms in MONTH_WEEK_MAP.items()})
    return jsonify(ok=False, error="データのパースに失敗"), 400


@app.route("/api/week-mapping")
def api_week_mapping_get():
    """現在の週マッピングを返す"""
    return jsonify(
        month_week_map={str(y): {str(m): wks for m, wks in ms.items()}
                        for y, ms in MONTH_WEEK_MAP.items()},
    )


try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(timezone="Asia/Tokyo")
    _scheduler.add_job(
        auto_archive_current_week,
        trigger="cron",
        day_of_week="wed",  # 会議(火曜)の翌日にアーカイブ
        hour=10,
        minute=0,
        id="auto_archive_weekly",
        replace_existing=True,
    )
    # 週マッピング: 毎週月曜 06:00 に GAS から更新
    def _scheduled_week_mapping_refresh():
        import week_mapping as _wm
        gas_url = _wm.GAS_WEEK_URL or os.environ.get("GAS_WEEK_URL", "")
        if gas_url:
            _refresh_week_from_gas(db=db, gas_url=gas_url)
        else:
            print("[INFO] week_mapping: GAS_WEEK_URL 未設定、定期更新スキップ")
    _scheduler.add_job(
        _scheduled_week_mapping_refresh,
        trigger="cron",
        day_of_week="mon",
        hour=6,
        minute=0,
        id="week_mapping_refresh",
        replace_existing=True,
    )
    _scheduler.start()
    print("スケジューラー: 毎週木曜10:00 自動アーカイブ / 毎週月曜06:00 週マッピング更新 有効")
except Exception as _sched_e:
    print(f"スケジューラー起動エラー: {_sched_e}")


if __name__ == "__main__":
    print("=" * 50)
    print("会議資料ダッシュボード 起動中")
    print("http://localhost:5050")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5050, debug=False)
